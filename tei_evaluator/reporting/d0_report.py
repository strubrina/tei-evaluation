"""
Dimension 0 Reporting Module.

This module provides specialized reporting functionality for Dimension 0
(XML Well-formedness) evaluation results.
"""

import csv
import json
from pathlib import Path
from typing import Any, Dict, List

from ..models import EvaluationResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class D0Reporter:
    """
    Specialized reporting for Dimension 0 (XML Well-formedness) evaluation results.

    This reporter generates comprehensive reports in multiple formats (JSON, Excel, CSV)
    for XML well-formedness evaluation results.
    """

    def save_detailed_report(self, results: List[EvaluationResult], output_dir: str = "output"):
        """
        Save comprehensive reports in multiple formats.

        JSON is saved under a 'json' subfolder of output_dir.
        Excel report is saved directly under output_dir.

        Args:
            results: List of EvaluationResult objects
            output_dir: Output directory for reports (default: "output")
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        json_dir = output_path / "json"
        json_dir.mkdir(exist_ok=True)

        # Save JSON and Excel reports
        self.save_json_report(results, json_dir / "d0_wellformedness_report.json")
        self.save_excel_summary(results, output_path / "d0_wellformedness_report.xlsx")

    def save_json_report(self, results: List[EvaluationResult], file_path: Path):
        """
        Save detailed JSON report for wellformedness.

        Args:
            results: List of EvaluationResult objects
            file_path: Path to output JSON file
        """
        # Generate summary statistics
        summary = self._generate_summary(results)

        report_data = {
            "evaluation_dimension": 0,
            "evaluation_type": "wellformedness",
            "total_files": len(results),
            "summary": summary,
            "files": []
        }

        for result in results:
            error_breakdown = result.metrics.get('error_breakdown', {})

            # Filter out non-XML errors when calculating xml_errors
            from ..models import ErrorType
            xml_errors_count = len([e for e in result.errors if e.type not in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]])

            # Similar fix for critical_errors
            critical_errors_count = len([e for e in result.errors if e.severity >= 8 and e.type not in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]])

            # Collect non-XML errors (e.g., FILE_NOT_FOUND, CONTENT_PRESERVATION)
            non_xml_errors = [{
                    "type": error.type.value,
                    "location": error.location,
                    "message": error.message
                } for error in result.errors
                if error.type in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]]

            file_not_found = any(e.type == ErrorType.FILE_NOT_FOUND for e in result.errors) or result.metrics.get('xml_file_missing', False)

            file_data = {
                "file_name": result.metrics.get('file_name', ''),
                "well_formed": result.metrics.get('well_formed', False),
                "xml_score": result.metrics.get('xml_score', result.score),
                "xml_errors": xml_errors_count,
                "critical_errors": critical_errors_count,
                "file_not_found": file_not_found,
                "non_xml_errors": non_xml_errors,
                "xml_malformed": error_breakdown.get('xml_malformed', 0),
                "character_encoding": error_breakdown.get('character_encoding', 0),
                "tag_structure": error_breakdown.get('tag_structure', 0),
                "attribute_syntax": error_breakdown.get('attribute_syntax', 0),
                "detailed_analysis": result.metrics.get('detailed_analysis', {}),
                "errors": [{
                        "type": error.type.value,
                        "location": error.location,
                        "message": error.message
                    } for error in result.errors
                    if error.type not in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]]
            }

            report_data["files"].append(file_data)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    def save_excel_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save Excel summary for wellformedness.

        Args:
            results: List of EvaluationResult objects
            file_path: Path to output Excel file
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            csv_path = file_path.with_suffix('.csv')
            self.save_csv_summary(results, csv_path)
            return

        wb = Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Sheet 1: Summary statistics
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary = self._generate_summary(results)

        # Write summary statistics
        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['B1'].alignment = Alignment(horizontal="right", vertical="center")

        summary_data = [
            ("Total Files", summary['total_files']),
            ("Well-formed Files", summary['well_formed_files']),
            ("Well-formed Rate (%)", f"{summary['well_formed_rate']:.2f}"),
            ("Average XML Score", f"{summary.get('average_xml_score', 0):.2f}"),
            ("Min XML Score", f"{summary.get('min_xml_score', 0):.2f}"),
            ("Max XML Score", f"{summary.get('max_xml_score', 0):.2f}"),
            ("Files Not Found", summary.get('files_not_found', 0)),
            ("Non-XML Errors", summary.get('non_xml_errors_total', 0)),
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 2):
            ws_summary[f'A{row_idx}'] = metric
            ws_summary[f'B{row_idx}'] = value
            ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # Sheet 2: Detailed results
        ws = wb.create_sheet(title="Detailed Results")
        # Headers for wellformedness only
        headers = [
            'File Name', 'Well Formed', 'XML Score',
            'XML Errors', 'Critical Errors',
            'XML Malformed', 'Character Encoding', 'Tag Structure', 'Attribute Syntax'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows for wellformedness only
        for row_idx, result in enumerate(results, 2):
            error_breakdown = result.metrics.get('error_breakdown', {})
            # Recompute counts with JSON-style exclusions to ensure parity
            from ..models import ErrorType
            xml_errors_count = len([e for e in result.errors if e.type not in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]])
            critical_errors_count = len([e for e in result.errors if e.severity >= 8 and e.type not in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]])
            xml_score_value = result.metrics.get('xml_score', result.score)
            row_data = [
                result.metrics.get('file_name', ''),
                result.metrics.get('well_formed', False),
                xml_score_value,
                xml_errors_count,
                critical_errors_count,
                error_breakdown.get('xml_malformed', 0),
                error_breakdown.get('character_encoding', 0),
                error_breakdown.get('tag_structure', 0),
                error_breakdown.get('attribute_syntax', 0)
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-adjust column widths for detailed results sheet
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass  # Cell value cannot be converted to string
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add conditional formatting for boolean columns and scores
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Apply formatting
        for row_idx in range(2, len(results) + 2):
            # Well Formed column (column B)
            cell = ws.cell(row=row_idx, column=2)
            if cell.value == True or cell.value == "TRUE":
                cell.fill = green_fill
            elif cell.value == False or cell.value == "FALSE":
                cell.fill = red_fill

            # XML Score column (column C) - color code by percentage
            cell = ws.cell(row=row_idx, column=3)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 90:
                    cell.fill = green_fill
                elif cell.value >= 70:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

        # Sheet 3: XML Error Details (only for files with XML errors)
        files_with_xml_errors = [r for r in results if r.metrics.get('xml_errors', 0) > 0]

        if files_with_xml_errors:
            ws_xml_errors = wb.create_sheet(title="XML Error Details")

            # Headers for XML errors sheet
            xml_error_headers = [
                'File Name', 'Error #', 'Error Type', 'Location', 'Message'
            ]

            for col, header in enumerate(xml_error_headers, 1):
                cell = ws_xml_errors.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write XML errors
            error_row = 2
            for result in files_with_xml_errors:
                file_name = result.metrics.get('file_name', 'Unknown')
                # Get only XML-related errors (exclude file_not_found and content_preservation)
                from ..models import ErrorType
                xml_errors = [e for e in result.errors if e.type not in [ErrorType.CONTENT_PRESERVATION, ErrorType.FILE_NOT_FOUND]]

                for i, error in enumerate(xml_errors, 1):
                    ws_xml_errors.cell(row=error_row, column=1, value=file_name if i == 1 else "")
                    ws_xml_errors.cell(row=error_row, column=2, value=i)
                    ws_xml_errors.cell(row=error_row, column=3, value=error.type.value)
                    ws_xml_errors.cell(row=error_row, column=4, value=error.location)
                    ws_xml_errors.cell(row=error_row, column=5, value=error.message)
                    error_row += 1

            # Auto-adjust column widths for XML errors sheet
            ws_xml_errors.column_dimensions['A'].width = 20  # File Name
            ws_xml_errors.column_dimensions['B'].width = 8   # Error #
            ws_xml_errors.column_dimensions['C'].width = 20  # Error Type
            ws_xml_errors.column_dimensions['D'].width = 25  # Location
            ws_xml_errors.column_dimensions['E'].width = 60  # Message

        # Sheet 4: Non-XML Errors (e.g., file not found)
        from ..models import ErrorType
        files_with_nonxml_errors = [r for r in results if any(e.type in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION] for e in r.errors)]

        if files_with_nonxml_errors:
            ws_nonxml = wb.create_sheet(title="Non-XML Errors")

            nonxml_headers = [
                'File Name', 'Error #', 'Error Type', 'Location', 'Message'
            ]

            for col, header in enumerate(nonxml_headers, 1):
                cell = ws_nonxml.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            error_row = 2
            for result in files_with_nonxml_errors:
                file_name = result.metrics.get('file_name', 'Unknown')
                nonxml_errors = [e for e in result.errors if e.type in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION]]

                for i, error in enumerate(nonxml_errors, 1):
                    ws_nonxml.cell(row=error_row, column=1, value=file_name if i == 1 else "")
                    ws_nonxml.cell(row=error_row, column=2, value=i)
                    ws_nonxml.cell(row=error_row, column=3, value=error.type.value)
                    ws_nonxml.cell(row=error_row, column=4, value=error.location)
                    ws_nonxml.cell(row=error_row, column=5, value=error.message)
                    error_row += 1

            ws_nonxml.column_dimensions['A'].width = 20
            ws_nonxml.column_dimensions['B'].width = 8
            ws_nonxml.column_dimensions['C'].width = 20
            ws_nonxml.column_dimensions['D'].width = 25
            ws_nonxml.column_dimensions['E'].width = 60

        wb.save(file_path)

    def save_csv_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save CSV summary (fallback method).

        Args:
            results: List of EvaluationResult objects
            file_path: Path to output CSV file
        """
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Headers for wellformedness only
            headers = [
                'File Name', 'Well Formed', 'XML Score',
                'XML Errors', 'Critical Errors',
                'XML Malformed', 'Character Encoding', 'Tag Structure', 'Attribute Syntax'
            ]
            writer.writerow(headers)

            # Data rows for wellformedness only
            for result in results:
                error_breakdown = result.metrics.get('error_breakdown', {})
                row = [
                    result.metrics.get('file_name', ''),
                    result.metrics.get('well_formed', False),
                    result.score,
                    result.metrics.get('xml_errors', len(result.errors)),
                    result.metrics.get('critical_errors', 0),
                    error_breakdown.get('xml_malformed', 0),
                    error_breakdown.get('character_encoding', 0),
                    error_breakdown.get('tag_structure', 0),
                    error_breakdown.get('attribute_syntax', 0)
                ]
                writer.writerow(row)

    def _generate_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """
        Generate summary statistics for wellformedness.

        Args:
            results: List of EvaluationResult objects

        Returns:
            Dictionary containing summary statistics
        """
        if not results:
            return {}

        total_files = len(results)
        passed_files = sum(1 for r in results if r.passed)
        failed_files = total_files - passed_files
        pass_rate = (passed_files / total_files * 100) if total_files > 0 else 0

        # Count total errors
        total_errors = sum(len(r.errors) for r in results)

        # Generate error breakdown
        all_errors = []
        for result in results:
            all_errors.extend(result.errors)

        error_breakdown = {}
        for error in all_errors:
            error_type = error.type.value
            error_breakdown[error_type] = error_breakdown.get(error_type, 0) + 1

        # Filter error breakdown for wellformedness only
        dimension0_error_types = {
            'xml_malformed',
            'character_encoding',
            'tag_structure',
            'attribute_syntax'
        }
        filtered_error_breakdown = {k: v for k, v in error_breakdown.items() if k in dimension0_error_types and v > 0}

        # XML-related summary data with scores
        xml_scores = [r.metrics.get('xml_score', r.score) for r in results]
        well_formed_files = sum(1 for r in results if r.metrics.get('well_formed', False))

        # Non-XML errors statistics
        from ..models import ErrorType
        files_not_found = sum(1 for r in results if any(e.type == ErrorType.FILE_NOT_FOUND for e in r.errors) or r.metrics.get('xml_file_missing', False))
        non_xml_errors_total = sum(1 for e in all_errors if e.type in [ErrorType.FILE_NOT_FOUND, ErrorType.CONTENT_PRESERVATION])

        return {
            "total_files": total_files,
            "passed_files": passed_files,
            "failed_files": failed_files,
            "pass_rate": pass_rate,
            "average_xml_score": sum(xml_scores) / len(xml_scores) if xml_scores else 0,
            "min_xml_score": min(xml_scores) if xml_scores else 0,
            "max_xml_score": max(xml_scores) if xml_scores else 0,
            "well_formed_files": well_formed_files,
            "well_formed_rate": (well_formed_files / total_files * 100) if total_files > 0 else 0,
            "total_errors": total_errors,
            "error_breakdown": filtered_error_breakdown,
            "files_not_found": files_not_found,
            "non_xml_errors_total": non_xml_errors_total
        }

    def generate_batch_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """
        Generate summary statistics for batch evaluation results.

        Args:
            results: List of EvaluationResult objects

        Returns:
            Dictionary containing batch summary statistics
        """
        if not results:
            return {"total_files": 0, "message": "No results to summarize"}

        total_files = len(results)
        passed_files = sum(1 for r in results if r.passed)
        failed_files = total_files - passed_files

        xml_scores = [r.metrics.get('xml_score', r.score) for r in results]
        avg_xml_score = sum(xml_scores) / len(xml_scores) if xml_scores else 0

        # Error type breakdown
        all_errors = []
        for result in results:
            all_errors.extend(result.errors)

        from ..models import ErrorType
        error_breakdown = {}
        for error_type in ErrorType:
            count = len([e for e in all_errors if e.type == error_type])
            if count > 0:
                error_breakdown[error_type.value] = count

        # Well-formedness stats
        well_formed_count = sum(1 for r in results if r.metrics.get('well_formed', False))

        summary = {
            "total_files": total_files,
            "passed_files": passed_files,
            "failed_files": failed_files,
            "pass_rate": (passed_files / total_files * 100) if total_files > 0 else 0,
            "average_xml_score": avg_xml_score,
            "min_xml_score": min(xml_scores) if xml_scores else 0,
            "max_xml_score": max(xml_scores) if xml_scores else 0,
            "well_formed_files": well_formed_count,
            "well_formed_rate": (well_formed_count / total_files * 100) if total_files > 0 else 0,
            "total_errors": len(all_errors),
            "error_breakdown": error_breakdown,
            "detailed_results": [
                {
                    "file_name": result.metrics.get('file_name', f"File_{i+1}"),
                    "file_path": result.metrics.get('file_path', ''),
                    "passed": result.passed,
                    "xml_score": result.metrics.get('xml_score', result.score),
                    "well_formed": result.metrics.get('well_formed', False),
                    "xml_errors": result.metrics.get('xml_errors', 0),
                    "error_count": len(result.errors)
                }
                for i, result in enumerate(results)
            ]
        }

        return summary

    def print_batch_summary(self, results: List[EvaluationResult]):
        """
        Print a formatted summary of batch evaluation results.

        Args:
            results: List of EvaluationResult objects
        """
        summary = self.generate_batch_summary(results)

        print("\n" + "="*60)
        print("DIMENSION 0 WELLFORMEDNESS EVALUATION SUMMARY")
        print("="*60)
        print(f"Total Files: {summary['total_files']}")
        print(f"Passed: {summary['passed_files']} ({summary['pass_rate']:.1f}%)")
        print(f"Failed: {summary['failed_files']}")

        print(f"\nWELLFORMEDNESS")
        print(f"Average XML Score: {summary.get('average_xml_score', 0):.2f}/100")
        print(f"XML Score Range: {summary.get('min_xml_score', 0):.2f} - {summary.get('max_xml_score', 0):.2f}")
        print(f"Well-formed: {summary['well_formed_files']}/{summary['total_files']} ({summary['well_formed_rate']:.1f}%)")
        print(f"Total XML Errors: {summary['total_errors']}")

        print("\nFile Details:")
        for detail in summary['detailed_results']:
            status = "[VALID]" if detail['passed'] else "[INVALID]"
            print(f"  {status} {detail['file_name']:<25} XML: {detail['xml_score']:6.2f} | Errors: {detail['xml_errors']}")

        print("="*60)


