"""
Dimension 3: Structural Comparison Reporting.

This module provides specialized reporting functionality for Dimension 3 structural
evaluation results, including JSON, Excel, and CSV output formats.
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from ..models import EvaluationResult


class CustomJSONEncoder(json.JSONEncoder):
    """Custom JSON encoder to handle non-serializable objects"""
    def default(self, obj):
        if isinstance(obj, bool):
            return obj
        try:
            return super().default(obj)
        except TypeError:
            return str(obj)

# Excel support with fallback to CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class D3Reporter:
    """
    Specialized reporting for Dimension 3 structural evaluation results.

    Generates comprehensive reports in multiple formats (JSON, Excel, CSV) with
    detailed structural analysis including completeness and XMLDiff operations.
    """

    def save_detailed_report(self, results: List[EvaluationResult], output_dir: str = "output"):
        """
        Save comprehensive reports in multiple formats.

        JSON is saved under a 'json' subfolder of output_dir.
        Excel report is saved directly under output_dir.

        Args:
            results: List of evaluation results to report
            output_dir: Directory to save reports (default: "output")
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        json_dir = output_path / "json"
        json_dir.mkdir(exist_ok=True)

        # Save JSON report
        self.save_json_report(results, json_dir / "d3_structural_fidelity_report.json")

        # Save Excel report (with CSV fallback)
        self.save_excel_summary(results, output_path / "d3_structural_fidelity_report.xlsx")

        print(f"[OUTPUT] Dimension 3 reports saved to: {output_path} (JSON in {json_dir})")

    def _generate_error_breakdown(self, errors):
        """Generate error breakdown without requiring ErrorType import"""
        breakdown = {}
        for error in errors:
            error_type = error.type.value if hasattr(error.type, 'value') else str(error.type)
            breakdown[error_type] = breakdown.get(error_type, 0) + 1
        return breakdown

    def save_json_report(self, results: List[EvaluationResult], file_path: Path):
        """
        Save detailed JSON report with comprehensive structural analysis.

        Args:
            results: List of evaluation results
            file_path: Path to save JSON report
        """
        report_data = {
            "timestamp": datetime.now().isoformat(),
            "evaluation_dimension": 3,
            "evaluation_type": "Structural Comparison Against Reference",
            "summary": self._generate_summary(results),
            "files": []
        }

        for result in results:
            # Extract analysis data (same logic as Excel)
            detailed_analysis = result.metrics.get('detailed_analysis', {})
            summary = detailed_analysis.get('summary', {}) if detailed_analysis else {}
            completeness_metrics = summary.get('completeness_metrics', result.metrics.get('completeness_metrics', {})) if summary else result.metrics.get('completeness_metrics', {})

            comparison_details = result.metrics.get('comparison_details', {})
            quick_check = comparison_details.get('quick_check', {})

            # Check comparison status
            is_structural_match = result.metrics.get('structural_match', False)
            comparison_skipped = result.metrics.get('comparison_skipped', False)

            # Element counts - prioritize quick_check data (total element counts)
            xml_elements = quick_check.get('xml_element_count', 0)
            ref_elements = quick_check.get('ref_element_count', 0)
            element_diff = xml_elements - ref_elements

            # Calculate elements added/removed from element_count_changes
            element_count_changes = summary.get('element_count_changes', {})
            elements_added = 0
            elements_removed = 0

            for change_data in element_count_changes.values():
                diff = change_data.get('difference', 0)
                if diff > 0:
                    elements_added += diff
                elif diff < 0:
                    elements_removed += abs(diff)

            # Determine analysis availability
            analysis_available = summary.get('analysis_available', False) or result.metrics.get('analysis_available', False)

            # Completeness Match: TRUE only when analysis ran and counts match
            completeness_match = (analysis_available and
                                  xml_elements == ref_elements and
                                  elements_added == 0 and
                                  elements_removed == 0)

            # XMLDiff data and tree edit distance
            xmldiff_data = comparison_details.get('xmldiff', {})
            tree_edit_distance = result.metrics.get('tree_edit_distance', None)
            normalized_ted = result.metrics.get('tree_edit_normalized', None)
            tree_similarity = result.metrics.get('tree_similarity', None)
            tree_lcs_similarity = result.metrics.get('tree_lcs_similarity', None)

            if tree_similarity is None and isinstance(normalized_ted, (int, float)):
                tree_similarity = max(0.0, 1.0 - normalized_ted)
            if tree_lcs_similarity is None and isinstance(tree_similarity, (int, float)):
                tree_lcs_similarity = tree_similarity

            if comparison_skipped:
                xmldiff_ops = None
                xmldiff_structural_ops = None
                xmldiff_text_ops = None
                tree_edit_value = None
                similarity_value = None
                lcs_value = None
            elif is_structural_match:
                xmldiff_ops = xmldiff_data.get('total_differences', 0)
                xmldiff_structural_ops = xmldiff_data.get('structural_differences', 0)
                xmldiff_text_ops = xmldiff_data.get('text_differences', 0)
                tree_edit_value = tree_edit_distance if tree_edit_distance is not None else 0
                similarity_value = tree_similarity if tree_similarity is not None else 1.0
                lcs_value = tree_lcs_similarity if tree_lcs_similarity is not None else 1.0
            else:
                xmldiff_ops = xmldiff_data.get('total_differences')
                xmldiff_structural_ops = xmldiff_data.get('structural_differences')
                xmldiff_text_ops = xmldiff_data.get('text_differences')
                tree_edit_value = tree_edit_distance
                similarity_value = tree_similarity
                lcs_value = tree_lcs_similarity

            if isinstance(tree_edit_value, (int, float)):
                tree_edit_display = round(float(tree_edit_value), 2)
            else:
                tree_edit_display = tree_edit_value
            if isinstance(similarity_value, (int, float)):
                similarity_display = round(float(similarity_value), 2)
            else:
                similarity_display = similarity_value
            if isinstance(lcs_value, (int, float)):
                # Scale LCS from 0-1 to 0-100 range for consistency with other scores
                lcs_display = round(float(lcs_value) * 100, 2)
            else:
                lcs_display = lcs_value

            # Align JSON with Excel: include xml_body_found and xmldiff_score
            xml_body_found = summary.get('xml_body_found', True)
            if comparison_skipped:
                xmldiff_score_value = None
            elif is_structural_match:
                xmldiff_score_value = result.metrics.get('xmldiff_score', 100.0)
            else:
                xmldiff_score_value = result.metrics.get('xmldiff_score', None)
            xmldiff_score_display = round(float(xmldiff_score_value), 1) if isinstance(xmldiff_score_value, (int, float)) else xmldiff_score_value

            file_data = {
                "file_name": result.metrics.get('file_name', ''),
                "reference_found": result.metrics.get('reference_found', False),
                "xml_body_found": xml_body_found,
                "completeness_match": completeness_match,
                "completeness_score": round(result.metrics.get('completeness_score', 0), 1),
                "tree_edit_distance": tree_edit_display,
                "ted_similarity": similarity_display,
                "tree_similarity": similarity_display,
                "tree_lcs_similarity": lcs_display,
                "xmldiff_score": xmldiff_score_display,
                "xml_elements": xml_elements,
                "ref_elements": ref_elements,
                "difference": element_diff,
                "elements_added": elements_added,
                "elements_removed": elements_removed,
                "completeness_recall": round(completeness_metrics.get('recall', 0), 1),
                "completeness_precision": round(completeness_metrics.get('precision', 0), 1),
                "completeness_f1": round(completeness_metrics.get('f1', 0), 1),
                "completeness_macro_f1": round(completeness_metrics.get('macro_f1', 0), 1),
                "completeness_micro_f1": round(completeness_metrics.get('micro_f1', 0), 1),
                "structural_match": is_structural_match,
                "structural_score": round(result.metrics.get('structural_score', 0), 1),
                "xmldiff_operations": xmldiff_ops,
                "xmldiff_structural_ops": xmldiff_structural_ops,
                "xmldiff_text_ops": xmldiff_text_ops,
                "errors": [
                    {
                        "type": error.type.value,
                        "location": error.location,
                        "message": error.message
                    } for error in result.errors
                ]
            }

            report_data["files"].append(file_data)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False, cls=CustomJSONEncoder)

    def save_excel_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save Excel summary with 4 sheets.

        Sheets: Summary, Detailed Results, XMLDiff Operations, Completeness Analysis

        Args:
            results: List of evaluation results
            file_path: Path to save Excel file
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            csv_file = file_path.with_suffix('.csv')
            return self.save_csv_summary(results, csv_file)

        wb = Workbook()

        # Define styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Generate summary statistics
        summary = self._generate_summary(results)

        # SHEET 1: Summary Statistics
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['B1'].alignment = Alignment(horizontal="right", vertical="center")

        summary_data = [
            ("Total Files", summary['total_files']),
            ("Skipped (No Reference)", summary['skipped_files']),
            ("", ""),  # Blank row
            ("Structural Matches", summary['matched_files']),
            ("Structural Differences", summary['diff_files']),
            ("Avg Structural Score (%)", f"{summary.get('avg_structural_score', 0):.1f}"),
            ("Min Structural Score (%)", f"{summary.get('min_structural_score', 0):.1f}"),
            ("Max Structural Score (%)", f"{summary.get('max_structural_score', 0):.1f}")
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 2):
            ws_summary[f'A{row_idx}'] = metric
            ws_summary[f'B{row_idx}'] = value
            ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # SHEET 2: Detailed Results
        ws = wb.create_sheet(title="Detailed Results")

        # Define headers - UPDATED ORDER
        headers = [
            'File Name', 'Reference Found', 'Body Found',
            'Structural Match', 'Structural Score', 'Tree LCS Similarity (%)', 'Tree Edit Distance', 'TED Similarity',
            'XMLDiff Operations', 'XMLDiff Score',
            'Completeness Match', 'Completeness Score',
            'LLM Elements', 'Ref Elements', 'Difference', 'Elements Added', 'Elements Removed',
            'Recall (%)', 'Precision (%)', 'F1 (%)', 'Macro F1 (%)', 'Micro F1 (%)'
        ]

        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_idx, result in enumerate(results, 2):  # Start from row 2
            # Extract analysis data
            detailed_analysis = result.metrics.get('detailed_analysis', {})
            summary = detailed_analysis.get('summary', {}) if detailed_analysis else {}
            completeness_metrics = (
                summary.get('completeness_metrics', result.metrics.get('completeness_metrics', {}))
                if summary else result.metrics.get('completeness_metrics', {})
            )

            comparison_details = result.metrics.get('comparison_details', {})
            quick_check = comparison_details.get('quick_check', {})

            # Check if this is a structural match
            is_structural_match = result.metrics.get('structural_match', False)
            comparison_skipped = result.metrics.get('comparison_skipped', False)

            # Element counts - prioritize quick_check data (total element counts)
            xml_elements = quick_check.get('xml_element_count', 0)
            ref_elements = quick_check.get('ref_element_count', 0)
            element_diff = xml_elements - ref_elements

            # Calculate elements added/removed from element_count_changes
            element_count_changes = summary.get('element_count_changes', {})
            elements_added = 0
            elements_removed = 0

            for change_data in element_count_changes.values():
                diff = change_data.get('difference', 0)
                if diff > 0:
                    elements_added += diff
                elif diff < 0:
                    elements_removed += abs(diff)

            # Determine analysis availability
            analysis_available = summary.get('analysis_available', False) or result.metrics.get('analysis_available', False)

            # Completeness Match: only TRUE when analysis ran and counts match
            completeness_match = (analysis_available and
                                  xml_elements == ref_elements and
                                  elements_added == 0 and
                                  elements_removed == 0)

            # Get scores
            completeness_score = result.metrics.get('completeness_score', 0)
            structural_score = result.metrics.get('structural_score', 0)
            completeness_recall = completeness_metrics.get('recall', result.metrics.get('completeness_recall', 0))
            completeness_precision = completeness_metrics.get('precision', result.metrics.get('completeness_precision', 0))
            completeness_f1 = completeness_metrics.get('f1', completeness_score)
            completeness_macro_f1 = completeness_metrics.get('macro_f1', result.metrics.get('completeness_macro_f1', 0))
            completeness_micro_f1 = completeness_metrics.get('micro_f1', result.metrics.get('completeness_micro_f1', completeness_score))

            # XMLDiff data
            xmldiff_data = comparison_details.get('xmldiff', {})
            tree_edit_distance = result.metrics.get('tree_edit_distance', None)
            normalized_ted = result.metrics.get('tree_edit_normalized', None)
            tree_similarity = result.metrics.get('tree_similarity', None)
            tree_lcs_similarity = result.metrics.get('tree_lcs_similarity', None)
            if tree_similarity is None and isinstance(normalized_ted, (int, float)):
                tree_similarity = max(0.0, 1.0 - normalized_ted)
            if tree_lcs_similarity is None and isinstance(tree_similarity, (int, float)):
                tree_lcs_similarity = tree_similarity

            if comparison_skipped:
                xmldiff_ops = 'N/A'
                tree_edit_value = 'N/A'
                similarity_value = 'N/A'
                lcs_value = 'N/A'
                xmldiff_score_value = 'N/A'
            elif is_structural_match:
                xmldiff_ops = xmldiff_data.get('total_differences', 0)
                tree_edit_value = tree_edit_distance if tree_edit_distance is not None else 0
                similarity_value = tree_similarity if tree_similarity is not None else 1.0
                lcs_value = tree_lcs_similarity if tree_lcs_similarity is not None else 1.0
                xmldiff_score_value = result.metrics.get('xmldiff_score', 100.0)
            else:
                xmldiff_ops = xmldiff_data.get('total_differences', 'N/A')
                tree_edit_value = tree_edit_distance if tree_edit_distance is not None else 'N/A'
                similarity_value = tree_similarity if tree_similarity is not None else 'N/A'
                lcs_value = tree_lcs_similarity if tree_lcs_similarity is not None else 'N/A'
                xmldiff_score_value = result.metrics.get('xmldiff_score', 'N/A')

            # Check if body was found in XML
            xml_body_found = summary.get('xml_body_found', True)  # Default to True for backward compatibility

            # Write row data - UPDATED COLUMN ORDER
            if isinstance(tree_edit_value, (int, float)):
                tree_edit_display = round(tree_edit_value, 2) if isinstance(tree_edit_value, float) else tree_edit_value
            else:
                tree_edit_display = tree_edit_value
            if isinstance(similarity_value, (int, float)):
                similarity_display = round(similarity_value, 2)
            else:
                similarity_display = similarity_value
            if isinstance(lcs_value, (int, float)):
                lcs_display = round(lcs_value, 2)
            else:
                lcs_display = lcs_value

            row_data = [
                result.metrics.get('file_name', ''),
                result.metrics.get('reference_found', False),
                xml_body_found,
                is_structural_match,
                round(structural_score, 1),  # Structural Score (now LCS-based %)
                lcs_display,
                tree_edit_display,
                similarity_display,
                xmldiff_ops,
                round(xmldiff_score_value, 1) if isinstance(xmldiff_score_value, (int, float)) else xmldiff_score_value,
                completeness_match,
                round(completeness_score, 1),  # Completeness Score (rounded to 1 decimal)
                xml_elements,  # Renamed header will display as LLM Elements
                ref_elements,
                element_diff,
                elements_added,
                elements_removed,
                round(completeness_recall, 1),
                round(completeness_precision, 1),
                round(completeness_f1, 1),
                round(completeness_macro_f1, 1),
                round(completeness_micro_f1, 1)
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass  # Cell value cannot be converted to string

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add conditional formatting for boolean columns and scores
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Apply formatting (updated column indices)
        for row_idx in range(2, len(results) + 2):

            # Structural Match (column 4)
            cell = ws.cell(row=row_idx, column=4)
            if cell.value is True:
                cell.fill = green_fill
            elif cell.value is False:
                cell.fill = red_fill

            # Structural Score (column 5) - color code by percentage
            cell = ws.cell(row=row_idx, column=5)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 90:
                    cell.fill = green_fill
                elif cell.value >= 70:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

            # Completeness Match (column 11)
            cell = ws.cell(row=row_idx, column=11)
            if cell.value is True:
                cell.fill = green_fill
            elif cell.value is False:
                cell.fill = red_fill

            # Completeness Score (column 12) - color code by percentage
            cell = ws.cell(row=row_idx, column=12)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 90:
                    cell.fill = green_fill
                elif cell.value >= 70:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

        # SHEET 3: Completeness Analysis (element type breakdown)
        ws_comp = wb.create_sheet(title="Completeness Analysis")

        # Headers
        comp_headers = ['File Name', 'Element Type', 'XML Count', 'Ref Count', 'Difference', 'Status']

        for col, header in enumerate(comp_headers, 1):
            cell = ws_comp.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows: element type breakdown for each file
        comp_row_idx = 2
        for result in results:
            detailed_analysis = result.metrics.get('detailed_analysis', {})
            summary = detailed_analysis.get('summary', {}) if detailed_analysis else {}
            element_count_changes = summary.get('element_count_changes', {})

            file_name = result.metrics.get('file_name', '')

            if element_count_changes:
                # Sort element types alphabetically
                sorted_types = sorted(element_count_changes.keys())

                # Track if this is the first row for this file (after filtering)
                first_row_for_file = True

                for elem_type in sorted_types:
                    change_data = element_count_changes[elem_type]
                    xml_count = change_data.get('xml_count', 0)
                    ref_count = change_data.get('ref_count', 0)
                    difference = change_data.get('difference', 0)

                    # Determine status
                    if ref_count == 0:
                        status = 'Added'
                    elif xml_count == 0:
                        status = 'Removed'
                    elif difference == 0:
                        status = 'Unchanged'
                    else:
                        status = 'Modified'

                    # Skip rows with "Unchanged" status
                    if status == 'Unchanged':
                        continue

                    # Write row (file name only on first row for this file)
                    ws_comp.cell(row=comp_row_idx, column=1, value=file_name if first_row_for_file else '')
                    ws_comp.cell(row=comp_row_idx, column=2, value=elem_type)
                    ws_comp.cell(row=comp_row_idx, column=3, value=xml_count)
                    ws_comp.cell(row=comp_row_idx, column=4, value=ref_count)
                    ws_comp.cell(row=comp_row_idx, column=5, value=difference)
                    ws_comp.cell(row=comp_row_idx, column=6, value=status)
                    comp_row_idx += 1
                    first_row_for_file = False
            else:
                # No element analysis available - determine why
                ref_found = result.metrics.get('reference_found', False)
                comparison_skipped = result.metrics.get('comparison_skipped', False)

                if not ref_found:
                    reason = 'Reference not found'
                elif comparison_skipped:
                    reason = 'Comparison skipped'
                else:
                    reason = 'Analysis unavailable'

                ws_comp.cell(row=comp_row_idx, column=1, value=file_name)
                ws_comp.cell(row=comp_row_idx, column=2, value='-')
                ws_comp.cell(row=comp_row_idx, column=3, value='-')
                ws_comp.cell(row=comp_row_idx, column=4, value='-')
                ws_comp.cell(row=comp_row_idx, column=5, value='-')
                ws_comp.cell(row=comp_row_idx, column=6, value=reason)
                comp_row_idx += 1

        # Set column widths for completeness sheet
        ws_comp.column_dimensions['A'].width = 25  # File Name
        ws_comp.column_dimensions['B'].width = 20  # Element Type
        ws_comp.column_dimensions['C'].width = 12  # XML Count
        ws_comp.column_dimensions['D'].width = 12  # Ref Count
        ws_comp.column_dimensions['E'].width = 12  # Difference
        ws_comp.column_dimensions['F'].width = 15  # Status

        # SHEET 4: XMLDiff Operation Breakdown (detailed list)
        ws_ops = wb.create_sheet(title="XMLDiff Operations")

        # Headers for detailed operation list
        headers = ['File Name', 'Operation #', 'Operation Type', 'Structural/Text', 'Description']

        for col, header in enumerate(headers, 1):
            cell = ws_ops.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows: list each individual operation
        row_idx = 2
        for result in results:
            comparison_details = result.metrics.get('comparison_details', {})
            xmldiff_data = comparison_details.get('xmldiff', {})

            # Get detailed operations
            detailed_ops = xmldiff_data.get('detailed_operations', [])

            if detailed_ops:
                file_name = result.metrics.get('file_name', '')
                for op_num, op in enumerate(detailed_ops, 1):
                    ws_ops.cell(row=row_idx, column=1, value=file_name if op_num == 1 else '')
                    ws_ops.cell(row=row_idx, column=2, value=op_num)
                    ws_ops.cell(row=row_idx, column=3, value=op.get('type', ''))
                    ws_ops.cell(row=row_idx, column=4, value='Structural' if op.get('is_structural', False) else 'Text')
                    ws_ops.cell(row=row_idx, column=5, value=op.get('description', ''))
                    row_idx += 1
            else:
                # No operations for this file (perfect match or no xmldiff data)
                ws_ops.cell(row=row_idx, column=1, value=result.metrics.get('file_name', ''))
                ws_ops.cell(row=row_idx, column=2, value=0)
                ws_ops.cell(row=row_idx, column=3, value='No operations')
                ws_ops.cell(row=row_idx, column=4, value='-')
                ws_ops.cell(row=row_idx, column=5, value='Perfect match or XMLDiff not available')
                row_idx += 1

        # Set column widths for operations sheet
        ws_ops.column_dimensions['A'].width = 25  # File Name
        ws_ops.column_dimensions['B'].width = 12  # Operation #
        ws_ops.column_dimensions['C'].width = 20  # Operation Type
        ws_ops.column_dimensions['D'].width = 15  # Structural/Text
        ws_ops.column_dimensions['E'].width = 80  # Description (element path + details)

        # Reorder sheets: place "XMLDiff Operations" before "Completeness Analysis"
        try:
            sheets = wb._sheets  # openpyxl internal list, stable across versions
            # Expected current order: ["Summary", "Detailed Results", "Completeness Analysis", "XMLDiff Operations"]
            # Desired order: ["Summary", "Detailed Results", "XMLDiff Operations", "Completeness Analysis"]
            if sheets[-2].title == "Completeness Analysis" and sheets[-1].title == "XMLDiff Operations":
                # Swap the last two sheets
                sheets[-2], sheets[-1] = sheets[-1], sheets[-2]
                wb._sheets = sheets
        except (AttributeError, IndexError, TypeError):
            # If reordering fails, continue without raising
            pass

        # Save the workbook
        wb.save(file_path)

    def save_csv_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save CSV summary for Dimension 3 structural evaluation (fallback method).

        Args:
            results: List of evaluation results
            file_path: Path to save CSV file
        """
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')

            # Headers - UPDATED ORDER matching Excel
            headers = [
                'File Name', 'Reference Found', 'Body Found',
                'Structural Match', 'Structural Score', 'Tree LCS Similarity (%)', 'Tree Edit Distance', 'TED Similarity',
                'XMLDiff Operations', 'XMLDiff Score',
                'Completeness Match', 'Completeness Score',
                'LLM Elements', 'Ref Elements', 'Difference', 'Elements Added', 'Elements Removed',
                'Recall (%)', 'Precision (%)', 'F1 (%)', 'Macro F1 (%)', 'Micro F1 (%)'
            ]
            writer.writerow(headers)

            # Data rows
            for result in results:
                # Extract analysis data
                detailed_analysis = result.metrics.get('detailed_analysis', {})
                summary = detailed_analysis.get('summary', {}) if detailed_analysis else {}
                completeness_metrics = summary.get('completeness_metrics', {}) if summary else {}

                comparison_details = result.metrics.get('comparison_details', {})
                quick_check = comparison_details.get('quick_check', {})

                # Check if this is a structural match
                is_structural_match = result.metrics.get('structural_match', False)
                comparison_skipped = result.metrics.get('comparison_skipped', False)

                # Element counts - prioritize quick_check data (total element counts)
                xml_elements = quick_check.get('xml_element_count', 0)
                ref_elements = quick_check.get('ref_element_count', 0)
                element_diff = xml_elements - ref_elements

                # Calculate elements added/removed from element_count_changes
                element_count_changes = summary.get('element_count_changes', {})
                elements_added = 0
                elements_removed = 0

                for change_data in element_count_changes.values():
                    diff = change_data.get('difference', 0)
                    if diff > 0:
                        elements_added += diff
                    elif diff < 0:
                        elements_removed += abs(diff)

                analysis_available = summary.get('analysis_available', False) or result.metrics.get('analysis_available', False)

                # Completeness Match: TRUE only when analysis ran and counts match
                completeness_match = (analysis_available and
                                     xml_elements == ref_elements and
                                     elements_added == 0 and
                                     elements_removed == 0)

                # Get scores
                completeness_score = result.metrics.get('completeness_score', 0)
                structural_score = result.metrics.get('structural_score', 0)
                completeness_recall = completeness_metrics.get('recall', result.metrics.get('completeness_recall', 0))
                completeness_precision = completeness_metrics.get('precision', result.metrics.get('completeness_precision', 0))
                completeness_f1 = completeness_metrics.get('f1', completeness_score)
                completeness_macro_f1 = completeness_metrics.get('macro_f1', result.metrics.get('completeness_macro_f1', 0))
                completeness_micro_f1 = completeness_metrics.get('micro_f1', result.metrics.get('completeness_micro_f1', completeness_score))

                # XMLDiff data
                xmldiff_data = comparison_details.get('xmldiff', {})
                tree_edit_distance = result.metrics.get('tree_edit_distance', None)
                normalized_ted = result.metrics.get('tree_edit_normalized', None)
                tree_similarity = result.metrics.get('tree_similarity', None)
                tree_lcs_similarity = result.metrics.get('tree_lcs_similarity', None)
                if tree_similarity is None and isinstance(normalized_ted, (int, float)):
                    tree_similarity = max(0.0, 1.0 - normalized_ted)
                if tree_lcs_similarity is None and isinstance(tree_similarity, (int, float)):
                    tree_lcs_similarity = tree_similarity

                if comparison_skipped:
                    xmldiff_ops = 'N/A'
                    tree_edit_value = 'N/A'
                    similarity_value = 'N/A'
                    lcs_value = 'N/A'
                    xmldiff_score_value = 'N/A'
                elif is_structural_match:
                    xmldiff_ops = xmldiff_data.get('total_differences', 0)
                    tree_edit_value = tree_edit_distance if tree_edit_distance is not None else 0
                    similarity_value = tree_similarity if tree_similarity is not None else 1.0
                    lcs_value = tree_lcs_similarity if tree_lcs_similarity is not None else 1.0
                    xmldiff_score_value = result.metrics.get('xmldiff_score', 100.0)
                else:
                    xmldiff_ops = xmldiff_data.get('total_differences', 'N/A')
                    tree_edit_value = tree_edit_distance if tree_edit_distance is not None else 'N/A'
                    similarity_value = tree_similarity if tree_similarity is not None else 'N/A'
                    lcs_value = tree_lcs_similarity if tree_lcs_similarity is not None else 'N/A'
                    xmldiff_score_value = result.metrics.get('xmldiff_score', 'N/A')

                # Write row data - UPDATED COLUMN ORDER
                if isinstance(tree_edit_value, (int, float)):
                    tree_edit_display = f"{float(tree_edit_value):.1f}"
                else:
                    tree_edit_display = tree_edit_value
                if isinstance(similarity_value, (int, float)):
                    similarity_display = f"{float(similarity_value):.1f}"
                else:
                    similarity_display = similarity_value
                if isinstance(lcs_value, (int, float)):
                    lcs_display = f"{float(lcs_value):.1f}"
                else:
                    lcs_display = lcs_value

                row_data = [
                    result.metrics.get('file_name', ''),
                    result.metrics.get('reference_found', False),
                    summary.get('xml_body_found', True),
                    is_structural_match,
                    f"{structural_score:.1f}",
                    lcs_display,
                    tree_edit_display,
                    similarity_display,
                    xmldiff_ops,
                    f"{xmldiff_score_value:.1f}" if isinstance(xmldiff_score_value, (int, float)) else xmldiff_score_value,
                    completeness_match,
                    f"{completeness_score:.1f}",
                    xml_elements,
                    ref_elements,
                    element_diff,
                    elements_added,
                    elements_removed,
                    f"{completeness_recall:.1f}",
                    f"{completeness_precision:.1f}",
                    f"{completeness_f1:.1f}",
                    f"{completeness_macro_f1:.1f}",
                    f"{completeness_micro_f1:.1f}"
                ]
                writer.writerow(row_data)

    def _generate_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """
        Generate summary statistics from evaluation results.

        Args:
            results: List of evaluation results

        Returns:
            Dictionary containing summary statistics
        """
        if not results:
            return {"total_files": 0}

        total_files = len(results)

        # Structural metrics
        matched_files = sum(1 for r in results if r.metrics.get('structural_match', False))
        diff_files = sum(1 for r in results if not r.metrics.get('structural_match', False) and not r.metrics.get('comparison_skipped', False))
        skipped_files = sum(1 for r in results if r.metrics.get('comparison_skipped', False))

        structural_scores = []

        for r in results:
            if not r.metrics.get('comparison_skipped', False):
                structural_scores.append(r.metrics.get('structural_score', 0))

        avg_structural = sum(structural_scores) / len(structural_scores) if structural_scores else 0

        return {
            "total_files": total_files,
            "matched_files": matched_files,
            "diff_files": diff_files,
            "skipped_files": skipped_files,
            "avg_structural_score": avg_structural,
            "min_structural_score": min(structural_scores) if structural_scores else 0,
            "max_structural_score": max(structural_scores) if structural_scores else 0,
            "match_rate": (matched_files / (total_files - skipped_files) * 100) if (total_files - skipped_files) > 0 else 0,
            "detailed_results": [
                {
                    "file_name": result.metrics.get('file_name', f"File_{i+1}"),
                    "structural_match": result.metrics.get('structural_match', False),
                    "comparison_skipped": result.metrics.get('comparison_skipped', False),
                    "score": round(result.score, 0),  # Keep overall score in detailed results
                    "error_count": len(result.errors)
                }
                for i, result in enumerate(results)
            ]
        }


