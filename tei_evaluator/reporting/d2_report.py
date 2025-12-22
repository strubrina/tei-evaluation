"""
Dimension 2: Schema Compliance Reporting.

This module provides specialized reporting functionality for Dimension 2 schema
evaluation results, including JSON, Excel, and CSV output formats.
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from ..models import EvaluationResult, ErrorType

# Excel support with fallback to CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class D2Reporter:
    """
    Specialized reporting for Dimension 2 schema evaluation results.

    Generates comprehensive reports in multiple formats (JSON, Excel, CSV) with
    support for different validation modes (TEI-only, project-only, combined).
    """

    def save_detailed_report(self, results: List[EvaluationResult], output_dir: str = "output"):
        """
        Save comprehensive reports in multiple formats.

        JSON is saved under a 'json' subfolder of output_dir.
        Excel report is saved directly under output_dir.

        Args:
            results: List of evaluation results to report
            output_dir: Directory to save reports (default: "output")
        """
        # Handle empty results
        if not results:
            print("[INFO] No results to save - skipping report generation")
            return

        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        json_dir = output_path / "json"
        json_dir.mkdir(exist_ok=True)

        # Detect evaluation mode from the results
        evaluation_mode = self._detect_evaluation_mode(results)

        if evaluation_mode == 'tei_validation_only':
            # TEI validation: JSON and Excel
            self.save_json_report(results, json_dir / "d2_tei_validation_report.json", evaluation_mode)
            self.save_excel_summary(results, output_path / "d2_tei_validation_report.xlsx", evaluation_mode)

        elif evaluation_mode == 'project_validation_only':
            # Project validation: JSON and Excel
            self.save_json_report(results, json_dir / "d2_project_validation_report.json", evaluation_mode)
            self.save_excel_summary(results, output_path / "d2_project_validation_report.xlsx", evaluation_mode)

        else:  # combined mode
            # Combined mode: JSON and Excel
            self.save_json_report(results, json_dir / "d2_schema_validation_report.json", evaluation_mode)
            self.save_excel_summary(results, output_path / "d2_schema_validation_report.xlsx", evaluation_mode)

        print(f"[OUTPUT] Dimension 2 reports saved to: {output_path} (JSON in {json_dir})")

    def _detect_evaluation_mode(self, results: List[EvaluationResult]) -> str:
        """
        Detect evaluation mode based on results.

        Args:
            results: List of evaluation results

        Returns:
            Evaluation mode: 'combined', 'tei_validation_only', or 'project_validation_only'
        """
        if not results:
            return 'combined'

        # Check the first result's metrics for evaluation_mode
        first_result = results[0]
        explicit_mode = first_result.metrics.get('evaluation_mode')
        if explicit_mode:
            return explicit_mode

        # Fallback: detect based on validation config
        validation_config = first_result.metrics.get('validation_config', {})
        tei_enabled = validation_config.get('tei_enabled', False)
        project_enabled = validation_config.get('project_enabled', False)

        if tei_enabled and project_enabled:
            return 'combined'
        elif tei_enabled and not project_enabled:
            return 'tei_validation_only'
        elif project_enabled and not tei_enabled:
            return 'project_validation_only'
        else:
            return 'combined'

    def _filter_error_breakdown(self, error_breakdown: Dict[str, int], evaluation_mode: str) -> Dict[str, int]:
        """Filter error breakdown to only include relevant error types for the evaluation mode"""
        if evaluation_mode == 'tei_validation_only':
            relevant_types = {'tei_schema_violation'}
        elif evaluation_mode == 'project_validation_only':
            relevant_types = {'project_schema_violation'}
        else:  # combined
            relevant_types = {'tei_schema_violation', 'project_schema_violation'}

        return {k: v for k, v in error_breakdown.items() if k in relevant_types and v > 0}

    def _is_malformed_xml(self, result: EvaluationResult) -> bool:
        """Check if XML is malformed (cannot be validated due to parse errors)"""
        return (
            result.metrics.get('xml_parse_error', False) or
            not result.metrics.get('file_readable', True)
        )

    def save_json_report(self, results: List[EvaluationResult], file_path: Path, evaluation_mode: str = 'combined'):
        """
        Save detailed JSON report based on evaluation mode.

        Args:
            results: List of evaluation results
            file_path: Path to save JSON report
            evaluation_mode: Evaluation mode ('combined', 'tei_validation_only', 'project_validation_only')
        """
        # Generate filtered summary based on evaluation mode
        filtered_summary = self._generate_filtered_summary(results, evaluation_mode)

        report_data = {
            "timestamp": datetime.now().isoformat(),
            "evaluation_dimension": 2,
            "evaluation_mode": evaluation_mode,
            "evaluation_type": "Schema Compliance and Standard Usage",
            "total_files": len(results),
            "summary": filtered_summary,
            "files": []
        }

        # Add section header based on evaluation mode
        if evaluation_mode == 'tei_validation_only':
            report_data["files"].append("TEI VALIDATION:")
        elif evaluation_mode == 'project_validation_only':
            report_data["files"].append("PROJECT VALIDATION:")
        elif evaluation_mode == 'combined':
            report_data["files"].append("COMBINED (TEI + PROJECT):")

        for result in results:
            # Filter error breakdown to remove irrelevant error types
            filtered_error_breakdown = self._filter_error_breakdown(
                result.metrics.get('error_breakdown', {}),
                evaluation_mode
            )

            if evaluation_mode == 'tei_validation_only':
                # Compute TEI valid status - "N/A" for malformed XML, otherwise based on score
                if self._is_malformed_xml(result):
                    tei_valid_computed = "N/A"
                else:
                    tei_valid_computed = result.score == 100.0

                file_data = {
                    "file_name": result.metrics.get('file_name', ''),
                    "tei_valid": tei_valid_computed,
                    # Include TEI schema violation count for parity with Excel
                    "tei_schema_violation": filtered_error_breakdown.get('tei_schema_violation', 0),
                    "tei_validity_reference": result.metrics.get('tei_validity_reference'),
                    "tei_validity_match": result.metrics.get('tei_validity_match'),
                    "errors": [
                        {
                            "type": error.type.value,
                            "location": error.location,
                            "message": error.message
                        } for error in result.errors
                    ]
                }

            elif evaluation_mode == 'project_validation_only':
                # Compute project valid status - "N/A" for malformed XML, otherwise based on score
                if self._is_malformed_xml(result):
                    project_valid_computed = "N/A"
                else:
                    project_valid_computed = result.score == 100.0

                file_data = {
                    "file_name": result.metrics.get('file_name', ''),
                    "project_valid": project_valid_computed,
                    "project_schema_violation": filtered_error_breakdown.get('project_schema_violation', 0),
                    "project_validity_reference": result.metrics.get('project_validity_reference'),
                    "project_validity_match": result.metrics.get('project_validity_match'),
                    "errors": [
                        {
                            "type": error.type.value,
                            "location": error.location,
                            "message": error.message
                        } for error in result.errors
                    ]
                }

            else:  # combined mode
                # Compute validity status - "N/A" for malformed XML, otherwise from metrics
                if self._is_malformed_xml(result):
                    tei_valid_computed = "N/A"
                    project_valid_computed = "N/A"
                else:
                    tei_valid_computed = result.metrics.get('tei_schema_valid', "N/A")
                    project_valid_computed = result.metrics.get('project_schema_valid', "N/A")

                # Get error breakdown
                error_breakdown = result.metrics.get('error_breakdown', {})

                file_data = {
                    "file_name": result.metrics.get('file_name', ''),
                    "file_path": result.metrics.get('file_path', ''),
                    "tei_valid": tei_valid_computed,
                    "tei_schema_violation": error_breakdown.get('tei_schema_violation', 0),
                    "tei_validity_reference": result.metrics.get('tei_validity_reference'),
                    "tei_validity_match": result.metrics.get('tei_validity_match'),
                    "project_valid": project_valid_computed,
                    "project_schema_violation": error_breakdown.get('project_schema_violation', 0),
                    "project_validity_reference": result.metrics.get('project_validity_reference'),
                    "project_validity_match": result.metrics.get('project_validity_match'),
                    "errors": [
                        {
                            "type": error.type.value,
                            "location": error.location,
                            "message": error.message
                        } for error in result.errors
                    ]
                }

            report_data["files"].append(file_data)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    def _generate_filtered_summary(self, results: List[EvaluationResult], evaluation_mode: str) -> Dict[str, Any]:
        """Generate mode-specific summary statistics"""
        if not results:
            return {}

        total_files = len(results)
        passed_files = sum(1 for r in results if r.passed)
        failed_files = total_files - passed_files
        pass_rate = (passed_files / total_files * 100) if total_files > 0 else 0

        # Count total errors
        total_errors = sum(len(r.errors) for r in results)

        # Generate error breakdown
        all_errors = []
        for result in results:
            all_errors.extend(result.errors)

        error_breakdown = {}
        for error in all_errors:
            error_type = error.type.value
            error_breakdown[error_type] = error_breakdown.get(error_type, 0) + 1

        # Filter error breakdown for evaluation mode
        filtered_error_breakdown = self._filter_error_breakdown(error_breakdown, evaluation_mode)

        # Exclude malformed XML files from base score calculations
        valid_results = [r for r in results if not self._is_malformed_xml(r)]
        malformed_files = [r for r in results if self._is_malformed_xml(r)]
        scores = [r.score for r in valid_results]

        base_summary = {
            "total_files": total_files,
            "passed_files": passed_files,
            "failed_files": failed_files,
            "pass_rate": pass_rate,
            "average_score": sum(scores) / len(scores) if scores else 0,
            "min_score": min(scores) if scores else 0,
            "max_score": max(scores) if scores else 0,
            "total_errors": total_errors,
            "error_breakdown": filtered_error_breakdown,
            "malformed_xml_files": len(malformed_files)
        }

        if evaluation_mode == 'tei_validation_only':
            # Only TEI-related summary data
            # Exclude malformed XML files from scores and valid counts
            valid_results = [r for r in results if not self._is_malformed_xml(r)]
            tei_scores = [r.score for r in valid_results]  # In TEI mode, score = tei score
            tei_valid_files = sum(1 for r in valid_results if r.score == 100.0)

            base_summary.update({
                "tei_enabled_files": total_files,
                "tei_valid_files": tei_valid_files,
                "tei_valid_rate": (tei_valid_files / total_files * 100) if total_files > 0 else 0,
                "average_tei_score": sum(tei_scores) / len(tei_scores) if tei_scores else 0,
                "min_tei_score": min(tei_scores) if tei_scores else 0,
                "max_tei_score": max(tei_scores) if tei_scores else 0
            })

        elif evaluation_mode == 'project_validation_only':
            # Only project-related summary data
            # Exclude malformed XML files from scores and valid counts
            valid_results = [r for r in results if not self._is_malformed_xml(r)]
            project_scores = [r.score for r in valid_results]  # In project mode, score = project score
            project_valid_files = sum(1 for r in valid_results if r.score == 100.0)

            base_summary.update({
                "project_enabled_files": total_files,
                "project_valid_files": project_valid_files,
                "project_valid_rate": (project_valid_files / total_files * 100) if total_files > 0 else 0,
                "average_project_score": sum(project_scores) / len(project_scores) if project_scores else 0,
                "min_project_score": min(project_scores) if project_scores else 0,
                "max_project_score": max(project_scores) if project_scores else 0
            })

        else:  # combined mode
            # Both TEI and project summary data
            # Exclude malformed XML files from scores and valid counts
            valid_results = [r for r in results if not self._is_malformed_xml(r)]
            tei_scores = [r.metrics.get('tei_score', 0) for r in valid_results]
            project_scores = [r.metrics.get('project_score', 0) for r in valid_results]
            tei_valid_files = sum(1 for r in valid_results if r.metrics.get('tei_schema_valid', False))
            project_valid_files = sum(1 for r in valid_results if r.metrics.get('project_schema_valid', False))

            base_summary.update({
                "tei_enabled_files": total_files,
                "tei_valid_files": tei_valid_files,
                "tei_valid_rate": (tei_valid_files / total_files * 100) if total_files > 0 else 0,
                "average_tei_score": sum(tei_scores) / len(tei_scores) if tei_scores else 0,
                "min_tei_score": min(tei_scores) if tei_scores else 0,
                "max_tei_score": max(tei_scores) if tei_scores else 0,
                "project_enabled_files": total_files,
                "project_valid_files": project_valid_files,
                "project_valid_rate": (project_valid_files / total_files * 100) if total_files > 0 else 0,
                "average_project_score": sum(project_scores) / len(project_scores) if project_scores else 0,
                "min_project_score": min(project_scores) if project_scores else 0,
                "max_project_score": max(project_scores) if project_scores else 0
            })

        return base_summary

    def save_excel_summary(self, results: List[EvaluationResult], file_path: Path, evaluation_mode: str = 'combined'):
        """
        Save Excel summary with 3 sheets: Summary, Detailed Results, Error Breakdown.

        Args:
            results: List of evaluation results
            file_path: Path to save Excel file
            evaluation_mode: Evaluation mode ('combined', 'tei_validation_only', 'project_validation_only')
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            csv_file = file_path.with_suffix('.csv')
            return self.save_csv_summary_fallback(results, csv_file, evaluation_mode)

        wb = Workbook()

        # Define styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Generate summary statistics
        summary = self._generate_filtered_summary(results, evaluation_mode)

        # SHEET 1: Summary Statistics
        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['B1'].alignment = Alignment(horizontal="right", vertical="center")

        if evaluation_mode == 'tei_validation_only':
            # Summary data for TEI validation only
            summary_data = [
                ("Total Files", summary['total_files']),
                ("TEI Valid", summary.get('tei_valid_files', 0)),
                ("TEI Validity Rate (%)", f"{summary.get('tei_valid_rate', 0):.2f}"),
            ]
            # Add malformed XML files count if any
            malformed_count = summary.get('malformed_xml_files', 0)
            if malformed_count > 0:
                summary_data.append(("Malformed XML files (N/A)", malformed_count))
            summary_data.append(("No Project validation (N/A)", ""))

            for row_idx, (metric, value) in enumerate(summary_data, 2):
                ws_summary[f'A{row_idx}'] = metric
                ws_summary[f'B{row_idx}'] = value
                ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

            # Adjust column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20

            # SHEET 2: Detailed Results
            ws_details = wb.create_sheet(title="Detailed Results")

            # Headers (File Path and TEI Score removed) + TEI Schema Violation
            headers = ['File Name', 'TEI Valid', 'TEI Schema Violation', 'TEI Validity Reference', 'TEI Validity Match']

            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data rows
            for row_idx, result in enumerate(results, 2):
                error_breakdown = result.metrics.get('error_breakdown', {})
                # Check for malformed XML - return "N/A" instead of False
                if self._is_malformed_xml(result):
                    tei_valid = "N/A"
                else:
                    tei_valid = result.score == 100.0
                row_data = [
                    result.metrics.get('file_name', ''),
                    tei_valid,
                    error_breakdown.get('tei_schema_violation', 0),
                    result.metrics.get('tei_validity_reference', None),
                    result.metrics.get('tei_validity_match', None)
                ]

                for col, value in enumerate(row_data, 1):
                    ws_details.cell(row=row_idx, column=col, value=value)

            # Auto-adjust column widths for detailed results
            for column in ws_details.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass  # Cell value cannot be converted to string
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width

            # Add conditional formatting for boolean columns and scores
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Apply formatting using header lookup (robust to column order)
            for row_idx in range(2, len(results) + 2):
                header_map = {ws_details.cell(row=1, column=c).value: c for c in range(1, ws_details.max_column + 1)}
                if 'TEI Valid' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['TEI Valid'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill
                if 'TEI Validity Match' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['TEI Validity Match'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill

            # SHEET 3: Error Breakdown
            ws_errors = wb.create_sheet(title="Error Breakdown")

            ws_errors['A1'] = "File Name"
            ws_errors['B1'] = "Error Type"
            ws_errors['C1'] = "Message"
            for col in ['A1', 'B1', 'C1']:
                ws_errors[col].font = header_font
                ws_errors[col].fill = header_fill
                ws_errors[col].alignment = header_alignment

            error_row = 2
            for result in results:
                file_name = result.metrics.get('file_name', '')
                # Show individual errors with details
                for error in result.errors:
                    if error.type.value in ['tei_schema_violation']:
                        ws_errors.cell(row=error_row, column=1, value=file_name)
                        ws_errors.cell(row=error_row, column=2, value=error.type.value)
                        ws_errors.cell(row=error_row, column=3, value=error.message)
                        error_row += 1

            ws_errors.column_dimensions['A'].width = 25
            ws_errors.column_dimensions['B'].width = 25
            ws_errors.column_dimensions['C'].width = 70

        elif evaluation_mode == 'project_validation_only':
            # Summary data for Project validation only
            summary_data = [
                ("Total Files", summary['total_files']),
                ("Project Valid", summary.get('project_valid_files', 0)),
                ("Project Validity Rate (%)", f"{summary.get('project_valid_rate', 0):.2f}"),
            ]
            # Add malformed XML files count if any
            malformed_count = summary.get('malformed_xml_files', 0)
            if malformed_count > 0:
                summary_data.append(("Malformed XML files (N/A)", malformed_count))
            summary_data.append(("No TEI validation (N/A)", ""))

            for row_idx, (metric, value) in enumerate(summary_data, 2):
                ws_summary[f'A{row_idx}'] = metric
                ws_summary[f'B{row_idx}'] = value
                ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20

            # SHEET 2: Detailed Results
            ws_details = wb.create_sheet(title="Detailed Results")

            # Headers (File Path and Project Score removed) + Project Schema Violation
            headers = ['File Name', 'Project Valid', 'Project Schema Violation', 'Project Validity Reference', 'Project Validity Match']

            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_idx, result in enumerate(results, 2):
                error_breakdown = result.metrics.get('error_breakdown', {})
                # Check for malformed XML - return "N/A" instead of False
                if self._is_malformed_xml(result):
                    project_valid = "N/A"
                else:
                    project_valid = result.score == 100.0
                row_data = [
                    result.metrics.get('file_name', ''),
                    project_valid,
                    error_breakdown.get('project_schema_violation', 0),
                    result.metrics.get('project_validity_reference', None),
                    result.metrics.get('project_validity_match', None),
                ]

                for col, value in enumerate(row_data, 1):
                    ws_details.cell(row=row_idx, column=col, value=value)

            for column in ws_details.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass  # Cell value cannot be converted to string
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width

            # Add conditional formatting for boolean columns and scores
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Apply formatting using header lookup (robust to column order)
            for row_idx in range(2, len(results) + 2):
                header_map = {ws_details.cell(row=1, column=c).value: c for c in range(1, ws_details.max_column + 1)}
                if 'Project Valid' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['Project Valid'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill
                if 'Project Validity Match' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['Project Validity Match'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill

            # SHEET 3: Error Breakdown
            ws_errors = wb.create_sheet(title="Error Breakdown")

            ws_errors['A1'] = "File Name"
            ws_errors['B1'] = "Error Type"
            ws_errors['C1'] = "Message"
            for col in ['A1', 'B1', 'C1']:
                ws_errors[col].font = header_font
                ws_errors[col].fill = header_fill
                ws_errors[col].alignment = header_alignment

            error_row = 2
            for result in results:
                file_name = result.metrics.get('file_name', '')
                # Show individual errors with details
                for error in result.errors:
                    if error.type.value in ['project_schema_violation']:
                        ws_errors.cell(row=error_row, column=1, value=file_name)
                        ws_errors.cell(row=error_row, column=2, value=error.type.value)
                        ws_errors.cell(row=error_row, column=3, value=error.message)
                        error_row += 1

            ws_errors.column_dimensions['A'].width = 25
            ws_errors.column_dimensions['B'].width = 25
            ws_errors.column_dimensions['C'].width = 70

        else:  # combined mode
            # Summary data for Combined validation with dynamic TEI/Project presence
            # Check for TEI/Project presence, excluding malformed XML files
            valid_results = [r for r in results if not self._is_malformed_xml(r)]
            has_tei = any('tei_schema_valid' in r.metrics for r in valid_results)
            has_project = any('project_schema_valid' in r.metrics for r in valid_results)

            summary_data = [("Total Files", summary['total_files'])]
            # Add malformed XML files count if any
            malformed_count = summary.get('malformed_xml_files', 0)
            if malformed_count > 0:
                summary_data.append(("Malformed XML files (N/A)", malformed_count))
            summary_data.append(("", ""))  # Empty row

            if has_tei:
                summary_data += [
                    ("TEI Valid", summary.get('tei_valid_files', 0)),
                    ("TEI Validity Rate (%)", f"{summary.get('tei_valid_rate', 0):.2f}")
                ]
            else:
                summary_data += [("No TEI validation (N/A)", "")]
            summary_data += [("", "")]  # Empty row
            if has_project:
                summary_data += [
                    ("Project Valid", summary.get('project_valid_files', 0)),
                    ("Project Validity Rate (%)", f"{summary.get('project_valid_rate', 0):.2f}")
                ]
            else:
                summary_data += [("No Project validation (N/A)", "")]

            for row_idx, (metric, value) in enumerate(summary_data, 2):
                ws_summary[f'A{row_idx}'] = metric
                ws_summary[f'B{row_idx}'] = value
                ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

            ws_summary.column_dimensions['A'].width = 30
            ws_summary.column_dimensions['B'].width = 20

            # SHEET 2: Detailed Results
            ws_details = wb.create_sheet(title="Detailed Results")

            # Dynamic headers depending on available validation data
            headers = ['File Name']
            if has_tei:
                headers += ['TEI Valid', 'TEI Schema Violation', 'TEI Validity Reference', 'TEI Validity Match']
            if has_project:
                headers += ['Project Valid', 'Project Schema Violation', 'Project Validity Reference', 'Project Validity Match']

            for col, header in enumerate(headers, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            for row_idx, result in enumerate(results, 2):
                error_breakdown = result.metrics.get('error_breakdown', {})
                row_data = [result.metrics.get('file_name', '')]
                if 'TEI Valid' in headers:
                    # Check for malformed XML - return "N/A" instead of False
                    if self._is_malformed_xml(result):
                        tei_valid_value = "N/A"
                    else:
                        tei_valid_value = result.metrics.get('tei_schema_valid', "N/A")
                    row_data += [
                        tei_valid_value,
                        error_breakdown.get('tei_schema_violation', 0),
                        result.metrics.get('tei_validity_reference', "N/A"),
                        result.metrics.get('tei_validity_match', "N/A"),
                    ]
                if 'Project Valid' in headers:
                    # Check for malformed XML - return "N/A" instead of False
                    if self._is_malformed_xml(result):
                        project_valid_value = "N/A"
                    else:
                        project_valid_value = result.metrics.get('project_schema_valid', "N/A")
                    row_data += [
                        project_valid_value,
                        error_breakdown.get('project_schema_violation', 0),
                        result.metrics.get('project_validity_reference', "N/A"),
                        result.metrics.get('project_validity_match', "N/A"),
                    ]

                for col, value in enumerate(row_data, 1):
                    ws_details.cell(row=row_idx, column=col, value=value)

            for column in ws_details.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (TypeError, AttributeError):
                        pass  # Cell value cannot be converted to string
                adjusted_width = min(max_length + 2, 50)
                ws_details.column_dimensions[column_letter].width = adjusted_width

            # Add conditional formatting for boolean columns and scores
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Apply formatting using header lookup
            for row_idx in range(2, len(results) + 2):
                header_map = {ws_details.cell(row=1, column=c).value: c for c in range(1, ws_details.max_column + 1)}
                if 'TEI Valid' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['TEI Valid'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill
                if 'TEI Validity Match' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['TEI Validity Match'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill
                if 'Project Valid' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['Project Valid'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill
                if 'Project Validity Match' in header_map:
                    cell = ws_details.cell(row=row_idx, column=header_map['Project Validity Match'])
                    if cell.value == True or cell.value == "TRUE":
                        cell.fill = green_fill
                    elif cell.value == False or cell.value == "FALSE":
                        cell.fill = red_fill

            # SHEET 3: Error Breakdown
            ws_errors = wb.create_sheet(title="Error Breakdown")

            ws_errors['A1'] = "File Name"
            ws_errors['B1'] = "Error Type"
            ws_errors['C1'] = "Message"
            for col in ['A1', 'B1', 'C1']:
                ws_errors[col].font = header_font
                ws_errors[col].fill = header_fill
                ws_errors[col].alignment = header_alignment

            error_row = 2
            for result in results:
                file_name = result.metrics.get('file_name', '')
                # Show individual errors with details
                for error in result.errors:
                    if error.type.value in ['tei_schema_violation', 'project_schema_violation']:
                        ws_errors.cell(row=error_row, column=1, value=file_name)
                        ws_errors.cell(row=error_row, column=2, value=error.type.value)
                        ws_errors.cell(row=error_row, column=3, value=error.message)
                        error_row += 1

            ws_errors.column_dimensions['A'].width = 25
            ws_errors.column_dimensions['B'].width = 25
            ws_errors.column_dimensions['C'].width = 70

        # Save the workbook
        wb.save(file_path)

    def save_csv_summary_fallback(self, results: List[EvaluationResult], file_path: Path, evaluation_mode: str = 'combined'):
        """Fallback CSV summary when openpyxl is not available"""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            if evaluation_mode == 'tei_validation_only':
                # Headers for TEI validation only
                headers = [
                    'File Name', 'File Path', 'TEI Valid', 'TEI Score',
                    'TEI Validity Reference', 'TEI Validity Match'
                ]
                writer.writerow(headers)

                # Data rows for TEI validation only
                for result in results:
                    error_breakdown = result.metrics.get('error_breakdown', {})
                    # TEI Valid based on score (100.0 = valid)
                    tei_valid = result.score == 100.0
                    row = [
                        result.metrics.get('file_name', ''),
                        result.metrics.get('file_path', ''),
                        tei_valid,
                        result.score,
                        result.metrics.get('tei_validity_reference'),
                        result.metrics.get('tei_validity_match'),
                    ]
                    writer.writerow(row)

            elif evaluation_mode == 'project_validation_only':
                # Headers for project validation only
                headers = [
                    'File Name', 'File Path', 'Project Valid', 'Project Score',
                    'Project Validity Reference', 'Project Validity Match'
                ]
                writer.writerow(headers)

                # Data rows for project validation only
                for result in results:
                    error_breakdown = result.metrics.get('error_breakdown', {})
                    # Project Valid based on score (100.0 = valid)
                    project_valid = result.score == 100.0
                    row = [
                        result.metrics.get('file_name', ''),
                        result.metrics.get('file_path', ''),
                        project_valid,
                        result.score,
                        result.metrics.get('project_validity_reference'),
                        result.metrics.get('project_validity_match'),
                    ]
                    writer.writerow(row)

            else:  # combined mode
                # Headers for combined analysis
                headers = [
                    'File Name', 'File Path',
                    'TEI Valid', 'TEI Score', 'TEI Schema Violation', 'TEI Validity Reference', 'TEI Validity Match',
                    'Project Valid', 'Project Score', 'Project Validity Reference', 'Project Validity Match'
                ]
                writer.writerow(headers)

                # Data rows for combined analysis
                for result in results:
                    error_breakdown = result.metrics.get('error_breakdown', {})
                    row = [
                        result.metrics.get('file_name', ''),
                        result.metrics.get('file_path', ''),
                        result.metrics.get('tei_schema_valid', False),
                        result.metrics.get('tei_score', 0),
                        error_breakdown.get('tei_schema_violation', 0),
                        result.metrics.get('tei_validity_reference'),
                        result.metrics.get('tei_validity_match'),
                        result.metrics.get('project_schema_valid', False),
                        result.metrics.get('project_score', 0),
                        result.metrics.get('project_validity_reference'),
                        result.metrics.get('project_validity_match'),
                    ]
                    writer.writerow(row)

    def save_tei_jing_categories_excel(self, results: List[EvaluationResult], file_path: Path):
        """Save TEI Jing categories analysis to Excel"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV
            csv_file = file_path.with_suffix('.csv')
            return self.save_tei_jing_categories_csv(results, csv_file)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "TEI Jing Categories"

        # Define styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ['File Name', 'Category', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        row_idx = 2
        for result in results:
            tei_validation = result.metrics.get('tei_validation', {})
            jing_categories = tei_validation.get('jing_categories', {})

            for category, count in jing_categories.items():
                ws.cell(row=row_idx, column=1, value=result.metrics.get('file_name', ''))
                ws.cell(row=row_idx, column=2, value=category)
                ws.cell(row=row_idx, column=3, value=count)
                row_idx += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def save_project_jing_categories_excel(self, results: List[EvaluationResult], file_path: Path):
        """Save Project Jing categories analysis to Excel"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV
            csv_file = file_path.with_suffix('.csv')
            return self.save_project_jing_categories_csv(results, csv_file)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Project Jing Categories"

        # Define styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ['File Name', 'Category', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        row_idx = 2
        for result in results:
            project_validation = result.metrics.get('project_validation', {})
            jing_categories = project_validation.get('jing_categories', {})

            for category, count in jing_categories.items():
                ws.cell(row=row_idx, column=1, value=result.metrics.get('file_name', ''))
                ws.cell(row=row_idx, column=2, value=category)
                ws.cell(row=row_idx, column=3, value=count)
                row_idx += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def save_tei_vs_project_analysis_excel(self, results: List[EvaluationResult], file_path: Path):
        """Save TEI vs Project analysis to Excel"""
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV
            csv_file = file_path.with_suffix('.csv')
            return self.save_tei_vs_project_analysis(results, csv_file)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "TEI vs Project Analysis"

        # Define styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            'File Name', 'TEI Score', 'Project Score', 'Score Difference',
            'TEI Valid', 'Project Valid', 'Category', 'TEI Errors', 'Project Errors', 'Primary Issue'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row_idx, result in enumerate(results, 2):
            tei_score = result.metrics.get('tei_score', 0)
            project_score = result.metrics.get('project_score', 0)
            score_diff = tei_score - project_score
            tei_valid = result.metrics.get('tei_schema_valid', False)
            project_valid = result.metrics.get('project_schema_valid', False)

            # Determine category
            if tei_valid and project_valid:
                category = "Both Valid"
            elif tei_valid and not project_valid:
                category = "TEI Only Valid"
            elif not tei_valid and project_valid:
                category = "Project Only Valid"
            else:
                category = "Both Invalid"

            # Determine primary issue
            tei_errors = result.metrics.get('tei_errors', 0)
            project_errors = result.metrics.get('project_errors', 0)

            if tei_errors > project_errors:
                primary_issue = "TEI Schema"
            elif project_errors > tei_errors:
                primary_issue = "Project Schema"
            else:
                primary_issue = "Equal Issues"

            row_data = [
                result.metrics.get('file_name', ''),
                tei_score,
                project_score,
                score_diff,
                tei_valid,
                project_valid,
                category,
                tei_errors,
                project_errors,
                primary_issue
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def save_separated_reports(self, results: List[EvaluationResult], base_output_dir: Path):
        """
        Save reports based on validation mode.

        For combined validation (TEI + Project):
        - d2_detailed_report.json - Combined report with all validation data (in json/)
        - d2_schema_validation_report.xlsx - Combined summary with all columns (root)

        Note: Individual TEI and project reports are NOT generated in combined mode
        as they would just duplicate information already in the combined report.
        """
        # Handle empty results
        if not results:
            print("[INFO] No results to save - skipping report generation")
            return

        base_output_dir.mkdir(parents=True, exist_ok=True)
        json_dir = base_output_dir / "json"
        json_dir.mkdir(exist_ok=True)

        # Detect evaluation mode
        evaluation_mode = self._detect_evaluation_mode(results)

        if evaluation_mode == 'tei_validation_only':
            # TEI-only mode: Create TEI-focused reports
            tei_results = self._create_tei_validation_view(results)
            self.save_json_report(tei_results, json_dir / "d2_tei_validation_report.json", evaluation_mode='tei_validation_only')
            self.save_excel_summary(tei_results, base_output_dir / "d2_tei_validation_report.xlsx", evaluation_mode='tei_validation_only')

            print(f"\n[OUTPUT] Dimension 2 (TEI validation) reports saved to: {base_output_dir}")
            print(f"   - json/d2_tei_validation_report.json")
            print(f"   - d2_tei_validation_report.xlsx")

        elif evaluation_mode == 'project_validation_only':
            # Project-only mode: Create project-focused reports
            project_results = self._create_project_validation_view(results)
            self.save_json_report(project_results, json_dir / "d2_project_validation_report.json", evaluation_mode='project_validation_only')
            self.save_excel_summary(project_results, base_output_dir / "d2_project_validation_report.xlsx", evaluation_mode='project_validation_only')

            print(f"\n[OUTPUT] Dimension 2 (Project validation) reports saved to: {base_output_dir}")
            print(f"   - json/d2_project_validation_report.json")
            print(f"   - d2_project_validation_report.xlsx")

        else:  # combined mode
            # Combined mode: Only save combined reports (no separate TEI/project files)
            self.save_json_report(results, json_dir / "d2_schema_validation_report.json", evaluation_mode='combined')
            self.save_excel_summary(results, base_output_dir / "d2_schema_validation_report.xlsx", evaluation_mode='combined')

            print(f"\n[OUTPUT] Dimension 2 (Combined validation) reports saved to: {base_output_dir}")
            print(f"   - json/d2_schema_validation_report.json")
            print(f"   - d2_schema_validation_report.xlsx")

    def _create_tei_validation_view(self, results: List[EvaluationResult]) -> List[EvaluationResult]:
        """Create a TEI-validation-only view from combined evaluation results"""
        from ..models import EvaluationResult

        tei_results = []
        for result in results:
            # Filter errors to only include TEI-related errors
            tei_errors = [e for e in result.errors if e.type.value in ['tei_schema_violation']]

            # Create clean metrics focused only on TEI validation
            tei_score = result.metrics.get('tei_score', result.score)
            tei_schema_valid = result.metrics.get('tei_schema_valid', False)

            tei_metrics = {
                'file_path': result.metrics.get('file_path', ''),
                'file_name': result.metrics.get('file_name', ''),
                'tei_score': tei_score,
                'total_errors': len(tei_errors),
                'tei_errors': len(tei_errors),
                'tei_schema_valid': tei_schema_valid,
                'tei_validation': result.metrics.get('tei_validation', {}),
                'validation_config': {
                    'tei_enabled': True,
                    'project_enabled': False
                },
                'validation_mode': 'tei_only',
                'error_breakdown': {
                    error_type: len([e for e in tei_errors if e.type.value == error_type])
                    for error_type in ['tei_schema_violation']
                },
                'evaluation_mode': 'tei_validation_only'
            }

            # Determine pass status based on TEI score and validation
            tei_passed = tei_score >= 95.0 or (tei_schema_valid and len(tei_errors) == 0)

            tei_focused_result = EvaluationResult(
                dimension=result.dimension,
                passed=tei_passed,
                score=tei_score,
                errors=tei_errors,
                metrics=tei_metrics
            )
            tei_results.append(tei_focused_result)

        return tei_results

    def _save_tei_validation_reports(self, results: List[EvaluationResult], output_dir: Path):
        """Save reports focused on TEI validation (moved from batch test) - DEPRECATED, use save_separated_reports instead"""
        output_dir.mkdir(parents=True, exist_ok=True)
        tei_results = self._create_tei_validation_view(results)
        self.save_detailed_report(tei_results, str(output_dir))

    def _create_project_validation_view(self, results: List[EvaluationResult]) -> List[EvaluationResult]:
        """Create a project-validation-only view from combined evaluation results"""
        from ..models import EvaluationResult

        project_results = []
        for result in results:
            if 'project_validation' not in result.metrics:
                continue

            # Filter errors to only include project validation errors
            project_errors = [e for e in result.errors if e.type.value == 'project_schema_violation']

            # Create clean metrics focused only on project validation
            project_validation_data = result.metrics.get('project_validation', {})
            project_score = result.metrics.get('project_score', 0)
            project_schema_valid = result.metrics.get('project_schema_valid', False)

            project_metrics = {
                'file_path': result.metrics.get('file_path', ''),
                'file_name': result.metrics.get('file_name', ''),
                'project_score': project_score,
                'total_errors': len(project_errors),
                'project_errors': len(project_errors),
                'project_schema_valid': project_schema_valid,
                'project_validation': project_validation_data,
                'validation_config': {
                    'tei_enabled': False,
                    'project_enabled': True
                },
                'validation_mode': 'project_only',
                'error_breakdown': {
                    'project_schema_violation': len(project_errors)
                },
                'evaluation_mode': 'project_validation_only'
            }

            # Determine pass status based on project score and validation
            project_passed = project_score >= 95.0 or (project_schema_valid and len(project_errors) == 0)

            project_focused_result = EvaluationResult(
                dimension=result.dimension,
                passed=project_passed,
                score=project_score,
                errors=project_errors,
                metrics=project_metrics
            )
            project_results.append(project_focused_result)

        return project_results

    def _save_project_validation_reports(self, results: List[EvaluationResult], output_dir: Path):
        """Save reports focused on project validation (moved from batch test) - DEPRECATED, use save_separated_reports instead"""
        output_dir.mkdir(parents=True, exist_ok=True)
        project_results = self._create_project_validation_view(results)
        if project_results:
            self.save_detailed_report(project_results, str(output_dir))

    def save_tei_jing_categories_csv(self, results: List[EvaluationResult], file_path: Path):
        """Save detailed TEI Jing error categories analysis"""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Collect all unique Jing categories
            all_tei_categories = set()
            for result in results:
                tei_cats = result.metrics.get('tei_validation', {}).get('jing_categories', {})
                all_tei_categories.update(tei_cats.keys())

            if not all_tei_categories:
                # No TEI validation data
                writer.writerow(['No TEI validation data available'])
                return

            # Headers
            headers = ['File Name', 'TEI Valid', 'TEI Total Errors'] + sorted(all_tei_categories)
            writer.writerow(headers)

            # Data rows
            for result in results:
                tei_validation = result.metrics.get('tei_validation', {})
                tei_cats = tei_validation.get('jing_categories', {})

                row = [
                    result.metrics.get('file_name', ''),
                    result.metrics.get('tei_schema_valid', False),
                    result.metrics.get('tei_errors', 0)
                ]

                # Add category counts
                for category in sorted(all_tei_categories):
                    row.append(tei_cats.get(category, 0))

                writer.writerow(row)

    def save_project_jing_categories_csv(self, results: List[EvaluationResult], file_path: Path):
        """Save detailed Project Jing error categories analysis"""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Collect all unique Jing categories
            all_project_categories = set()
            for result in results:
                project_cats = result.metrics.get('project_validation', {}).get('jing_categories', {})
                all_project_categories.update(project_cats.keys())

            if not all_project_categories:
                # No project validation data
                writer.writerow(['No project validation data available'])
                return

            # Headers
            headers = ['File Name', 'Project Valid', 'Project Total Errors'] + sorted(all_project_categories)
            writer.writerow(headers)

            # Data rows
            for result in results:
                project_validation = result.metrics.get('project_validation', {})
                project_cats = project_validation.get('jing_categories', {})

                row = [
                    result.metrics.get('file_name', ''),
                    result.metrics.get('project_schema_valid', False),
                    result.metrics.get('project_errors', 0)
                ]

                # Add category counts
                for category in sorted(all_project_categories):
                    row.append(project_cats.get(category, 0))

                writer.writerow(row)

    def save_tei_vs_project_analysis(self, results: List[EvaluationResult], file_path: Path):
        """Save separate analysis of TEI vs Project compliance"""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Headers for comparison analysis
            headers = [
                'File Name', 'TEI Score', 'Project Score', 'Score Difference',
                'TEI Valid', 'Project Valid', 'Compliance Category',
                'TEI Errors', 'Project Errors', 'Primary Issue'
            ]
            writer.writerow(headers)

            # Data rows with analysis
            for result in results:
                tei_score = result.metrics.get('tei_score', 0)
                project_score = result.metrics.get('project_score', 0)
                score_diff = project_score - tei_score

                tei_valid = result.metrics.get('tei_schema_valid', False)
                project_valid = result.metrics.get('project_schema_valid', False)

                # Categorize compliance
                if tei_valid and project_valid:
                    category = "Full Compliance"
                elif tei_valid and not project_valid:
                    category = "TEI Only"
                elif not tei_valid and project_valid:
                    category = "Project Only"
                else:
                    category = "Non-Compliant"

                # Identify primary issue
                tei_errors = result.metrics.get('tei_errors', 0)
                project_errors = result.metrics.get('project_errors', 0)

                if tei_errors > project_errors:
                    primary_issue = "TEI Issues"
                elif project_errors > tei_errors:
                    primary_issue = "Project Issues"
                elif tei_errors > 0:
                    primary_issue = "Both Equal"
                else:
                    primary_issue = "None"

                writer.writerow([
                    result.metrics.get('file_name', ''),
                    tei_score,
                    project_score,
                    score_diff,
                    tei_valid,
                    project_valid,
                    category,
                    tei_errors,
                    project_errors,
                    primary_issue
                ])

    def _generate_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """Generate summary statistics using D2Schema logic"""
        from ..core.d2_schema import D2Schema
        evaluator = D2Schema()
        return evaluator.generate_batch_summary(results)


