"""
Dimension 1 Reporting Module.

This module provides specialized reporting functionality for Dimension 1
(Source Fidelity) evaluation results.
"""

import csv
import json
from pathlib import Path
from typing import Any, Dict, List

from ..models import EvaluationResult

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class D1Reporter:
    """
    Specialized reporting for Dimension 1 (Source Fidelity) evaluation results.

    This reporter generates comprehensive reports in multiple formats (JSON, Excel, CSV)
    for source fidelity evaluation results.
    """

    def save_detailed_report(self, results: List[EvaluationResult], output_dir: str = "output"):
        """
        Save comprehensive reports in multiple formats.

        JSON is saved under a 'json' subfolder of output_dir.
        Excel report is saved directly under output_dir.

        Args:
            results: List of EvaluationResult objects
            output_dir: Output directory for reports (default: "output")
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        json_dir = output_path / "json"
        json_dir.mkdir(exist_ok=True)

        # Save JSON and Excel reports
        self.save_json_report(results, json_dir / "d1_source_fidelity_report.json")
        self.save_excel_summary(results, output_path / "d1_source_fidelity_report.xlsx")

    def save_json_report(self, results: List[EvaluationResult], file_path: Path):
        """
        Save detailed JSON report for source fidelity.

        Args:
            results: List of EvaluationResult objects
            file_path: Path to output JSON file
        """
        # Generate summary statistics
        summary = self._generate_summary(results)

        report_data = {
            "evaluation_dimension": 1,
            "evaluation_type": "source_fidelity",
            "total_files": len(results),
            "summary": summary,
            "files": []
        }

        for result in results:
            content_preservation = result.metrics.get('content_preservation', {})
            diff_info = content_preservation.get('diff_info', [])
            diff_count = len(diff_info) if diff_info else 0

            # Prepare differences array mirroring Excel Differences sheet (raw content)
            differences = []
            if diff_info:
                for diff in diff_info:
                    differences.append({
                        "type": diff.get('type', ''),
                        "position": diff.get('position', 0),
                        "original_text": diff.get('original_text', ''),
                        "extracted_text": diff.get('extracted_text', ''),
                        "context": diff.get('context', '')
                    })

            file_data = {
                "file_name": result.metrics.get('file_name', ''),
                "content_preserved": result.metrics.get('content_score', 0) >= 100.0,  # Match Excel: >= 100%
                "content_score": round(result.metrics.get('content_score', 0), 2),
                "exact_match_no_whitespace": content_preservation.get('exact_match_without_whitespace', False),
                "exact_match_with_whitespace": content_preservation.get('exact_match_with_whitespace', False),
                "similarity_no_whitespace": round(content_preservation.get('similarity_without_whitespace', 0) * 100, 2),
                "similarity_with_whitespace": round(content_preservation.get('similarity_with_whitespace', 0) * 100, 2),
                "differences_count": diff_count,
                "differences": differences,
                "errors": [
                    {
                        "type": error.type.value,
                        "location": error.location,
                        "message": error.message
                    } for error in result.errors
                ]
            }

            report_data["files"].append(file_data)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    def save_excel_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save detailed source fidelity Excel report.

        Args:
            results: List of EvaluationResult objects
            file_path: Path to output Excel file
        """
        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            csv_path = file_path.with_suffix('.csv')
            self.save_csv_summary(results, csv_path)
            return

        wb = Workbook()

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Sheet 1: Summary statistics
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary = self._generate_summary(results)

        # Write summary statistics
        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['B1'].alignment = Alignment(horizontal="right", vertical="center")

        # Calculate content preservation rate
        content_preservation_rate = (summary['exact_content_matches'] / summary['total_files'] * 100) if summary['total_files'] > 0 else 0

        summary_data = [
            ("Total Files", summary['total_files']),
            ("Exact Matches", summary['exact_content_matches']),
            ("Source Fidelity Rate (%)", f"{content_preservation_rate:.2f}"),
            ("Average Content Score", f"{summary['average_content_score']:.2f}"),
            ("Min Content Score", f"{summary['min_content_score']:.2f}"),
            ("Max Content Score", f"{summary['max_content_score']:.2f}")
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 2):
            ws_summary[f'A{row_idx}'] = metric
            ws_summary[f'B{row_idx}'] = value
            ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

        # Adjust column widths
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 20

        # Sheet 2: Detailed results
        ws = wb.create_sheet(title="Detailed Results")

        # Headers
        headers = [
            'File Name',
            'Content Preserved', 'Content Score',
            'Exact Match (No Whitespace)', 'Exact Match (With Whitespace)',
            'Similarity (No Whitespace)', 'Similarity (With Whitespace)',
            'Differences Count'
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        row_idx = 2
        for result in results:
            content_preservation = result.metrics.get('content_preservation', {})
            if not content_preservation:
                continue  # Skip files without content analysis

            diff_info = content_preservation.get('diff_info', [])
            diff_count = len(diff_info) if diff_info else 0

            row_data = [
                result.metrics.get('file_name', ''),
                result.metrics.get('content_score', 0) >= 100.0,  # Content Preserved (TRUE if >= 100%)
                round(result.metrics.get('content_score', 0), 2),  # Content Score with 2 decimals
                content_preservation.get('exact_match_without_whitespace', False),
                content_preservation.get('exact_match_with_whitespace', False),
                round(content_preservation.get('similarity_without_whitespace', 0) * 100, 2),  # Convert to percentage with 2 decimals
                round(content_preservation.get('similarity_with_whitespace', 0) * 100, 2),  # Convert to percentage with 2 decimals
                diff_count
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)
            row_idx += 1

        # Auto-adjust column widths for detailed results
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass  # Cell value cannot be converted to string
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add conditional formatting for boolean columns and scores
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        # Apply formatting
        final_row = row_idx  # Save the final row_idx value before reusing it
        for row_idx in range(2, final_row):  # row_idx was incremented in the data loop

            # Content Preserved column (column B)
            cell = ws.cell(row=row_idx, column=2)  # Content Preserved
            if cell.value == True or cell.value == "TRUE":
                cell.fill = green_fill
            elif cell.value == False or cell.value == "FALSE":
                cell.fill = red_fill

            # Content Score column (column C) - color code by percentage
            cell = ws.cell(row=row_idx, column=3)  # Content Score
            if isinstance(cell.value, (int, float)):
                if cell.value >= 90:
                    cell.fill = green_fill
                elif cell.value >= 70:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

        # Sheet 3: Differences (only for files with differences)
        files_with_diffs = [r for r in results
                           if r.metrics.get('content_preservation', {}).get('diff_info')]

        if files_with_diffs:
            ws_diff = wb.create_sheet(title="Differences")

            # Headers for differences sheet
            diff_headers = [
                'File Name', 'Diff #', 'Type', 'Position',
                'Original Text', 'Extracted Text', 'Context'
            ]

            for col, header in enumerate(diff_headers, 1):
                cell = ws_diff.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Write differences
            diff_row = 2
            for result in files_with_diffs:
                content_pres = result.metrics.get('content_preservation', {})
                diff_info = content_pres.get('diff_info', [])
                file_name = result.metrics.get('file_name', 'Unknown')

                for i, diff in enumerate(diff_info, 1):
                    # Make whitespace visible
                    orig_text = diff.get('original_text', '').replace('\n', '\\n').replace('\t', '\\t').replace(' ', '·')
                    extr_text = diff.get('extracted_text', '').replace('\n', '\\n').replace('\t', '\\t').replace(' ', '·')
                    context = diff.get('context', '').replace('\n', '\\n').replace('\t', '\\t')

                    # Limit text length for readability
                    if len(orig_text) > 100:
                        orig_text = orig_text[:100] + "..."
                    if len(extr_text) > 100:
                        extr_text = extr_text[:100] + "..."
                    if len(context) > 150:
                        context = context[:150] + "..."

                    ws_diff.cell(row=diff_row, column=1, value=file_name if i == 1 else "")
                    ws_diff.cell(row=diff_row, column=2, value=i)
                    ws_diff.cell(row=diff_row, column=3, value=diff.get('type', ''))
                    ws_diff.cell(row=diff_row, column=4, value=diff.get('position', 0))
                    ws_diff.cell(row=diff_row, column=5, value=orig_text)
                    ws_diff.cell(row=diff_row, column=6, value=extr_text)
                    ws_diff.cell(row=diff_row, column=7, value=context)
                    diff_row += 1

            # Auto-adjust column widths for differences sheet
            ws_diff.column_dimensions['A'].width = 20  # File Name
            ws_diff.column_dimensions['B'].width = 8   # Diff #
            ws_diff.column_dimensions['C'].width = 12  # Type
            ws_diff.column_dimensions['D'].width = 10  # Position
            ws_diff.column_dimensions['E'].width = 30  # Original
            ws_diff.column_dimensions['F'].width = 30  # Extracted
            ws_diff.column_dimensions['G'].width = 40  # Context

        wb.save(file_path)

    def save_csv_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save detailed source fidelity CSV (fallback method).

        Args:
            results: List of EvaluationResult objects
            file_path: Path to output CSV file
        """
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                'File Name',
                'Content Preserved', 'Content Score',
                'Exact Match (No Whitespace)', 'Exact Match (With Whitespace)',
                'Similarity (No Whitespace)', 'Similarity (With Whitespace)',
                'Differences Count'
            ])

            # Data rows
            for result in results:
                content_preservation = result.metrics.get('content_preservation', {})
                if not content_preservation:
                    continue  # Skip files without content analysis

                diff_info = content_preservation.get('diff_info', [])
                diff_count = len(diff_info) if diff_info else 0

                writer.writerow([
                    result.metrics.get('file_name', ''),
                    result.metrics.get('content_score', 0) >= 100.0,  # Content Preserved (TRUE if >= 100%)
                    round(result.metrics.get('content_score', 0), 2),  # Content Score with 2 decimals
                    content_preservation.get('exact_match_without_whitespace', False),
                    content_preservation.get('exact_match_with_whitespace', False),
                    round(content_preservation.get('similarity_without_whitespace', 0) * 100, 2),  # Convert to percentage with 2 decimals
                    round(content_preservation.get('similarity_with_whitespace', 0) * 100, 2),  # Convert to percentage with 2 decimals
                    diff_count
                ])

    def _generate_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """
        Generate summary statistics for source fidelity.

        Args:
            results: List of EvaluationResult objects

        Returns:
            Dictionary containing summary statistics
        """
        if not results:
            return {}

        total_files = len(results)
        passed_files = sum(1 for r in results if r.passed)
        failed_files = total_files - passed_files
        pass_rate = (passed_files / total_files * 100) if total_files > 0 else 0

        # Count total errors
        total_errors = sum(len(r.errors) for r in results)

        # Generate error breakdown
        all_errors = []
        for result in results:
            all_errors.extend(result.errors)

        error_breakdown = {}
        for error in all_errors:
            error_type = error.type.value
            error_breakdown[error_type] = error_breakdown.get(error_type, 0) + 1

        # Only content-related summary data
        content_scores = [r.score for r in results]  # In content mode, score = content score
        exact_matches = sum(1 for r in results
                           if r.metrics.get('content_preservation', {}).get('exact_match_without_whitespace', False))
        txt_files_found = sum(1 for r in results
                             if r.metrics.get('content_preservation', {}).get('original_file_found', False))

        # Calculate content-specific pass/fail rates
        content_passed_files = sum(1 for r in results if r.score >= 100.0)  # Content score >= 100%
        content_failed_files = total_files - content_passed_files
        content_pass_rate = (content_passed_files / total_files * 100) if total_files > 0 else 0

        return {
            "total_files": total_files,
            "passed_files": content_passed_files,
            "failed_files": content_failed_files,
            "pass_rate": content_pass_rate,
            "average_content_score": round(sum(content_scores) / len(content_scores), 2) if content_scores else 0,
            "min_content_score": round(min(content_scores), 2) if content_scores else 0,
            "max_content_score": round(max(content_scores), 2) if content_scores else 0,
            "exact_content_matches": exact_matches,
            "txt_files_found": txt_files_found,
            "total_errors": total_errors,
            "error_breakdown": error_breakdown
        }

    def generate_batch_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """
        Generate summary statistics for batch evaluation results.

        Args:
            results: List of EvaluationResult objects

        Returns:
            Dictionary containing batch summary statistics
        """
        if not results:
            return {"total_files": 0, "message": "No results to summarize"}

        total_files = len(results)
        passed_files = sum(1 for r in results if r.passed)
        failed_files = total_files - passed_files

        # Content preservation statistics
        content_scores = [r.metrics.get('content_score', r.score) for r in results]
        avg_content_score = sum(content_scores) / len(content_scores) if content_scores else 0

        # Error type breakdown
        all_errors = []
        for result in results:
            all_errors.extend(result.errors)

        from ..models import ErrorType
        error_breakdown = {}
        for error_type in ErrorType:
            count = len([e for e in all_errors if e.type == error_type])
            if count > 0:
                error_breakdown[error_type.value] = count

        # Content preservation stats
        exact_matches = sum(1 for r in results
                        if r.metrics.get('content_preservation', {}).get('exact_match_without_whitespace', False))
        txt_files_found = sum(1 for r in results
                            if r.metrics.get('content_preservation', {}).get('original_file_found', False))

        summary = {
            "total_files": total_files,
            "passed_files": passed_files,
            "failed_files": failed_files,
            "pass_rate": (passed_files / total_files * 100) if total_files > 0 else 0,
            "average_content_score": round(avg_content_score, 2),
            "min_content_score": round(min(content_scores), 2) if content_scores else 0,
            "max_content_score": round(max(content_scores), 2) if content_scores else 0,
            "exact_content_matches": exact_matches,
            "txt_files_found": txt_files_found,
            "total_errors": len(all_errors),
            "error_breakdown": error_breakdown,
            "detailed_results": [
                {
                    "file_name": result.metrics.get('file_name', f"File_{i+1}"),
                    "file_path": result.metrics.get('file_path', ''),
                    "passed": result.passed,
                    "content_score": round(result.metrics.get('content_score', result.score), 2),
                    "content_errors": result.metrics.get('content_errors', 0),
                    "error_count": len(result.errors),
                    "content_match": result.metrics.get('content_preservation', {}).get('exact_match_without_whitespace', False)
                }
                for i, result in enumerate(results)
            ]
        }

        return summary

    def print_batch_summary(self, results: List[EvaluationResult]):
        """
        Print a formatted summary of batch evaluation results.

        Args:
            results: List of EvaluationResult objects
        """
        summary = self.generate_batch_summary(results)

        print("\n" + "="*60)
        print("DIMENSION 1 SOURCE FIDELITY EVALUATION SUMMARY")
        print("="*60)
        print(f"Total Files: {summary['total_files']}")
        print(f"Passed: {summary['passed_files']} ({summary['pass_rate']:.1f}%)")
        print(f"Failed: {summary['failed_files']}")

        print(f"\nSOURCE FIDELITY")
        print(f"Average Content Score: {summary['average_content_score']:.2f}/100")
        print(f"Content Score Range: {summary['min_content_score']:.2f} - {summary['max_content_score']:.2f}")
        print(f"Exact Matches: {summary['exact_content_matches']}/{summary['total_files']}")
        print(f"Total Content Errors: {summary['total_errors']}")

        print("\nFile Details:")
        for detail in summary['detailed_results']:
            status = "[PASS]" if detail['passed'] else "[FAIL]"
            content_status = "[EXACT]" if detail['content_match'] else "[DIFF]"
            print(f"  {status} {content_status} {detail['file_name']:<25} Content: {detail['content_score']:6.2f}")

        print("="*60)


