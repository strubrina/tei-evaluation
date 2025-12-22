"""
Dimension 4: Semantic Content Matching Reporting.

This module provides specialized reporting functionality for Dimension 4 semantic
evaluation results, including JSON, Excel, and CSV output formats with detailed
content matching analysis.
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any

from ..models import EvaluationResult, ErrorType

# Excel support with fallback to CSV
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class D4Reporter:
    """
    Specialized reporting for Dimension 4 semantic evaluation results.

    Generates comprehensive reports in multiple formats (JSON, Excel, CSV) with
    detailed content matching analysis including exact/practical match rates and
    element-level issues.
    """

    def save_detailed_report(self, results: List[EvaluationResult], output_dir: str = "output"):
        """
        Save comprehensive reports in multiple formats.

        JSON is saved under a 'json' subfolder of output_dir.
        Excel report is saved directly under output_dir.

        Args:
            results: List of evaluation results to report
            output_dir: Directory to save reports (default: "output")
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        json_dir = output_path / "json"
        json_dir.mkdir(exist_ok=True)

        # Save JSON report
        self.save_json_report(results, json_dir / "d4_semantic_validation_report.json")

        # Save Excel report (with CSV fallback)
        self.save_excel_summary(results, output_path / "d4_semantic_validation_report.xlsx")

        print(f"[OUTPUT] Dimension 4 reports saved to: {output_path} (JSON in {json_dir})")

    def save_json_report(self, results: List[EvaluationResult], file_path: Path):
        """
        Save detailed JSON report with comprehensive content matching analysis.

        Args:
            results: List of evaluation results
            file_path: Path to save JSON report
        """
        report_data = {
            "timestamp": datetime.now().isoformat(),
            "evaluation_dimension": 4,
            "evaluation_type": "Semantic Content Matching Against Reference Files",
            "summary": self._generate_summary(results),
            "files": []
        }

        for result in results:
            # Extract content matching data
            summary_stats = result.metrics.get('summary_stats', {})
            content = summary_stats.get('content', {})
            content_metrics = content.get('metrics', {})
            comparison_skipped = result.metrics.get('comparison_skipped', False)

            # Error breakdown by category
            error_breakdown = self._categorize_errors(result.errors)

            # Inclusion occurrences from comparison_results (more reliable than errors)
            inclusion_counts = self._count_inclusion_occurrences_from_metrics(result.metrics)

            # Calculate semantic match values
            semantic_match = result.passed

            # Determine relaxed semantic match (with safe int coercion)
            def _to_int_safe(v):
                if isinstance(v, int):
                    return v
                if isinstance(v, float):
                    try:
                        return int(v)
                    except (ValueError, TypeError):
                        return None
                if isinstance(v, str):
                    try:
                        return int(v.strip())
                    except (ValueError, TypeError):
                        return None
                return None

            total_elements = _to_int_safe(content.get('total_elements'))
            exact_matches = _to_int_safe(content.get('exact_matches'))
            practical_matches = _to_int_safe(content.get('practical_matches'))
            if isinstance(total_elements, int) and isinstance(exact_matches, int) and isinstance(practical_matches, int) and total_elements > 0:
                relaxed_semantic_match = (exact_matches + practical_matches == total_elements)
            else:
                relaxed_semantic_match = None

            # Format percentages (compute if nested rates missing but counts are available)
            exact_rate = content.get('exact_match_rate')
            practical_rate = content.get('practical_match_rate')
            if not isinstance(exact_rate, (int, float)):
                if isinstance(total_elements, int) and total_elements > 0 and isinstance(exact_matches, int):
                    exact_rate = (exact_matches / total_elements) * 100
            if not isinstance(practical_rate, (int, float)):
                if isinstance(total_elements, int) and total_elements > 0 and isinstance(practical_matches, int):
                    practical_rate = (practical_matches / total_elements) * 100

            # Create flat structured file data matching Excel columns
            file_data = {
                "file_name": result.metrics.get('file_name', ''),
                "reference_found": result.metrics.get('reference_found', False),
                "semantic_match": semantic_match,
                "relaxed_semantic_match": relaxed_semantic_match if isinstance(relaxed_semantic_match, bool) else "N/A",
                "score": round(result.score, 0),
                "compared_elements": total_elements if isinstance(total_elements, int) else "N/A",
                "missing_elements": content.get('missing_elements', "N/A"),
                "added_elements": content.get('extra_elements', "N/A"),
                "content_deviation": error_breakdown.get('content_errors', 0),
                "exact_matches": exact_matches if isinstance(exact_matches, int) else "N/A",
                "practical_matches": practical_matches if isinstance(practical_matches, int) else "N/A",
                "exact_match_rate": round(exact_rate, 0) if isinstance(exact_rate, (int, float)) else "N/A",
                "practical_match_rate": round(practical_rate, 0) if isinstance(practical_rate, (int, float)) else "N/A",
                "content_precision": round(content_metrics.get('precision'), 1) if isinstance(content_metrics.get('precision'), (int, float)) else "N/A",
                "content_recall": round(content_metrics.get('recall'), 1) if isinstance(content_metrics.get('recall'), (int, float)) else "N/A",
                "content_f1": round(content_metrics.get('f1'), 1) if isinstance(content_metrics.get('f1'), (int, float)) else "N/A",
                "content_macro_f1": round(content_metrics.get('macro_f1'), 1) if isinstance(content_metrics.get('macro_f1'), (int, float)) else "N/A",
                "content_micro_f1": round(content_metrics.get('micro_f1'), 1) if isinstance(content_metrics.get('micro_f1'), (int, float)) else "N/A",
                "over_inclusion_occurrences": inclusion_counts.get('over_inclusion_occurrences', error_breakdown.get('over_inclusion_occurrences', 0)),
                "under_inclusion_occurrences": inclusion_counts.get('under_inclusion_occurrences', error_breakdown.get('under_inclusion_occurrences', 0)),
                "errors": [
                    {
                        "type": error.type.value,
                        "location": error.location,
                        "message": error.message
                    } for error in result.errors
                ]
            }

            report_data["files"].append(file_data)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    def save_csv_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save CSV summary for Dimension 4 content matching evaluation (fallback method).

        Args:
            results: List of evaluation results
            file_path: Path to save CSV file
        """
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f, delimiter=';')

            # Header - aligned with JSON and Excel
            headers = [
                'File Name', 'Reference Found', 'Semantic Match', 'Relaxed Semantic Match', 'Score',
                'Compared Elements', 'Missing Elements', 'Added Elements', 'Content Deviation',
                'Exact Matches', 'Practical Matches',
                'Exact Match Rate (%)', 'Practical Match Rate (%)',
                'Content Precision (%)', 'Content Recall (%)', 'Content F1 (%)', 'Content Macro F1 (%)', 'Content Micro F1 (%)',
                'Over-inclusion Occurrences', 'Under-inclusion Occurrences'
            ]
            writer.writerow(headers)

            # Data rows
            for result in results:
                summary_stats = result.metrics.get('summary_stats', {})
                content = summary_stats.get('content', {})
                comparison_skipped = result.metrics.get('comparison_skipped', False)

                # Error breakdown by category
                error_breakdown = self._categorize_errors(result.errors)
                inclusion_counts = self._count_inclusion_occurrences_from_metrics(result.metrics)

                if comparison_skipped:
                    writer.writerow([
                        result.metrics.get('file_name', ''),
                        False,
                        'N/A', 'N/A', 'N/A',
                        'N/A', 'N/A', 'N/A', 'N/A',
                        'N/A', 'N/A',
                        'N/A', 'N/A',
                        'N/A', 'N/A', 'N/A', 'N/A', 'N/A',
                        0, 0
                    ])
                else:
                    # Coerce counts first
                    def _to_int_safe(v):
                        if isinstance(v, int):
                            return v
                        if isinstance(v, float):
                            try:
                                return int(v)
                            except (ValueError, TypeError):
                                return None
                        if isinstance(v, str):
                            try:
                                return int(v.strip())
                            except (ValueError, TypeError):
                                return None
                        return None
                    total_elements = _to_int_safe(content.get('total_elements'))
                    exact_matches = _to_int_safe(content.get('exact_matches'))
                    practical_matches = _to_int_safe(content.get('practical_matches'))
                    # Compute rates after counts available
                    exact_rate = content.get('exact_match_rate')
                    practical_rate = content.get('practical_match_rate')
                    if not isinstance(exact_rate, (int, float)):
                        if isinstance(total_elements, int) and total_elements > 0 and isinstance(exact_matches, int):
                            exact_rate = (exact_matches / total_elements) * 100
                    if not isinstance(practical_rate, (int, float)):
                        if isinstance(total_elements, int) and total_elements > 0 and isinstance(practical_matches, int):
                            practical_rate = (practical_matches / total_elements) * 100
                    content_metrics = content.get('metrics', {})

                    semantic_match = result.passed
                    if isinstance(total_elements, int) and isinstance(exact_matches, int) and isinstance(practical_matches, int) and total_elements > 0:
                        relaxed_semantic_match = (exact_matches + practical_matches == total_elements)
                    else:
                        relaxed_semantic_match = None

                    writer.writerow([
                        result.metrics.get('file_name', ''),
                        result.metrics.get('reference_found', False),
                        semantic_match,
                        relaxed_semantic_match if isinstance(relaxed_semantic_match, bool) else 'N/A',
                        f"{result.score:.0f}",
                        total_elements if isinstance(total_elements, int) else 'N/A',
                        content.get('missing_elements', 'N/A'),
                        content.get('extra_elements', 'N/A'),
                        error_breakdown.get('content_errors', 0),
                        exact_matches if isinstance(exact_matches, int) else 'N/A',
                        practical_matches if isinstance(practical_matches, int) else 'N/A',
                        f"{exact_rate:.0f}" if isinstance(exact_rate, (int, float)) else "N/A",
                        f"{practical_rate:.0f}" if isinstance(practical_rate, (int, float)) else "N/A",
                        f"{content_metrics.get('precision'):.1f}" if isinstance(content_metrics.get('precision'), (int, float)) else "N/A",
                        f"{content_metrics.get('recall'):.1f}" if isinstance(content_metrics.get('recall'), (int, float)) else "N/A",
                        f"{content_metrics.get('f1'):.1f}" if isinstance(content_metrics.get('f1'), (int, float)) else "N/A",
                        f"{content_metrics.get('macro_f1'):.1f}" if isinstance(content_metrics.get('macro_f1'), (int, float)) else "N/A",
                        f"{content_metrics.get('micro_f1'):.1f}" if isinstance(content_metrics.get('micro_f1'), (int, float)) else "N/A",
                        inclusion_counts.get('over_inclusion_occurrences', error_breakdown.get('over_inclusion_occurrences', 0)),
                        inclusion_counts.get('under_inclusion_occurrences', error_breakdown.get('under_inclusion_occurrences', 0))
                    ])

    def save_excel_summary(self, results: List[EvaluationResult], file_path: Path):
        """
        Save Excel summary for Dimension 4 content matching evaluation.

        Sheets: Summary, Detailed Results, Content Issues

        Args:
            results: List of evaluation results
            file_path: Path to save Excel file
        """
        if not OPENPYXL_AVAILABLE:
            csv_file = file_path.with_suffix('.csv')
            return self.save_csv_summary(results, csv_file)

        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        summary = self._generate_summary(results)

        ws_summary = wb.active
        ws_summary.title = "Summary"

        ws_summary['A1'] = "Metric"
        ws_summary['B1'] = "Value"
        ws_summary['A1'].font = header_font
        ws_summary['B1'].font = header_font
        ws_summary['A1'].fill = header_fill
        ws_summary['B1'].fill = header_fill
        ws_summary['B1'].alignment = Alignment(horizontal="right", vertical="center")

        summary_data = [
            ("Total Files", summary['total_files']),
            ("Skipped (No Reference)", summary['skipped_files']),
            ("Perfect File Matches", summary['perfect_file_matches']),
            ("", ""),
            ("--- BEST MATCH CONTENT MATCHING ---", ""),
            ("Elements Matched by Content", summary['content_stats']['total_comparisons']),
            ("Exact Content Matches", summary['content_stats']['exact_matches']),
            ("Content Exact Match Rate (%)", f"{summary['content_stats']['exact_match_rate']:.1f}"),
            ("Practical Content Matches", summary['content_stats']['practical_matches']),
            ("Content Practical Match Rate (%)", f"{summary['content_stats']['practical_match_rate']:.1f}"),
            ("Reordered Elements (matched but wrong position)", summary['content_stats']['reordered_elements']),
            ("", ""),
            ("Average Score", f"{summary['average_score']:.1f}"),
            ("Min Score", f"{summary['min_score']:.1f}"),
            ("Max Score", f"{summary['max_score']:.1f}")
        ]

        for row_idx, (metric, value) in enumerate(summary_data, 2):
            ws_summary[f'A{row_idx}'] = metric
            ws_summary[f'B{row_idx}'] = value
            ws_summary[f'B{row_idx}'].alignment = Alignment(horizontal="right", vertical="center")

        ws_summary.column_dimensions['A'].width = 45
        ws_summary.column_dimensions['B'].width = 20

        ws_details = wb.create_sheet(title="Detailed Results")

        headers = [
            'File Name', 'Reference Found', 'Semantic Match', 'Relaxed Semantic Match', 'Score',
            'Compared Elements', 'Missing Elements', 'Added Elements', 'Content Deviation','Exact Matches', 'Practical Matches',
            'Exact Match Rate (%)', 'Practical Match Rate (%)',
            'Content Precision (%)', 'Content Recall (%)', 'Content F1 (%)', 'Content Macro F1 (%)', 'Content Micro F1 (%)',
            'Over-inclusion Occurrences', 'Under-inclusion Occurrences'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        for row_idx, result in enumerate(results, 2):
            summary_stats = result.metrics.get('summary_stats', {})
            content = summary_stats.get('content', {})
            comparison_skipped = result.metrics.get('comparison_skipped', False)

            error_breakdown = self._categorize_errors(result.errors)
            inclusion_counts = self._count_inclusion_occurrences_from_metrics(result.metrics)

            if comparison_skipped:
                row_data = [
                    result.metrics.get('file_name', ''),
                    False,
                    'N/A',
                    'N/A',
                    'N/A',
                    'N/A', 'N/A', 'N/A', 'N/A',
                    'N/A', 'N/A',
                    'N/A', 'N/A',
                    'N/A', 'N/A', 'N/A', 'N/A', 'N/A',
                    0, 0
                ]
            else:
                # Coerce counts first
                def _to_int_safe(v):
                    if isinstance(v, int):
                        return v
                    if isinstance(v, float):
                        try:
                            return int(v)
                        except (ValueError, OverflowError):
                            return None
                    if isinstance(v, str):
                        try:
                            return int(v.strip())
                        except (ValueError, AttributeError):
                            return None
                    return None
                total_elements = _to_int_safe(content.get('total_elements'))
                exact_matches = _to_int_safe(content.get('exact_matches'))
                practical_matches = _to_int_safe(content.get('practical_matches'))
                # Compute rates after counts available
                exact_rate = content.get('exact_match_rate')
                practical_rate = content.get('practical_match_rate')
                if not isinstance(exact_rate, (int, float)):
                    if isinstance(total_elements, int) and total_elements > 0 and isinstance(exact_matches, int):
                        exact_rate = (exact_matches / total_elements) * 100
                if not isinstance(practical_rate, (int, float)):
                    if isinstance(total_elements, int) and total_elements > 0 and isinstance(practical_matches, int):
                        practical_rate = (practical_matches / total_elements) * 100
                content_metrics = content.get('metrics', {})

                semantic_match = result.passed
                if isinstance(total_elements, int) and isinstance(exact_matches, int) and isinstance(practical_matches, int) and total_elements > 0:
                    relaxed_semantic_match = (exact_matches + practical_matches == total_elements)
                else:
                    relaxed_semantic_match = None

                row_data = [
                    result.metrics.get('file_name', ''),
                    result.metrics.get('reference_found', False),
                    semantic_match,
                    relaxed_semantic_match if isinstance(relaxed_semantic_match, bool) else 'N/A',
                    round(result.score, 0),
                    total_elements if isinstance(total_elements, int) else 'N/A',
                    content.get('missing_elements', 'N/A'),
                    content.get('extra_elements', 'N/A'),
                    error_breakdown.get('content_errors', 0),
                    exact_matches if isinstance(exact_matches, int) else 'N/A',
                    practical_matches if isinstance(practical_matches, int) else 'N/A',
                    round(exact_rate, 0) if isinstance(exact_rate, (int, float)) else 'N/A',
                    round(practical_rate, 0) if isinstance(practical_rate, (int, float)) else 'N/A',
                    round(content_metrics.get('precision'), 1) if isinstance(content_metrics.get('precision'), (int, float)) else 'N/A',
                    round(content_metrics.get('recall'), 1) if isinstance(content_metrics.get('recall'), (int, float)) else 'N/A',
                    round(content_metrics.get('f1'), 1) if isinstance(content_metrics.get('f1'), (int, float)) else 'N/A',
                    round(content_metrics.get('macro_f1'), 1) if isinstance(content_metrics.get('macro_f1'), (int, float)) else 'N/A',
                    round(content_metrics.get('micro_f1'), 1) if isinstance(content_metrics.get('micro_f1'), (int, float)) else 'N/A',
                    inclusion_counts.get('over_inclusion_occurrences', error_breakdown.get('over_inclusion_occurrences', 0)),
                    inclusion_counts.get('under_inclusion_occurrences', error_breakdown.get('under_inclusion_occurrences', 0))
                ]

            for col, value in enumerate(row_data, 1):
                ws_details.cell(row=row_idx, column=col, value=value)

        for column in ws_details.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass  # Cell value cannot be converted to string
            adjusted_width = min(max_length + 2, 50)
            ws_details.column_dimensions[column_letter].width = adjusted_width

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for row_idx in range(2, len(results) + 2):
            cell = ws_details.cell(row=row_idx, column=3)
            if cell.value is True:
                cell.fill = green_fill
            elif cell.value is False:
                cell.fill = red_fill

            cell = ws_details.cell(row=row_idx, column=4)
            if cell.value is True:
                cell.fill = green_fill
            elif cell.value is False:
                cell.fill = red_fill

            cell = ws_details.cell(row=row_idx, column=5)
            if isinstance(cell.value, (int, float)):
                if cell.value >= 90:
                    cell.fill = green_fill
                elif cell.value >= 70:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill

        self._add_element_issues_sheet(wb, results, header_font, header_fill, header_alignment)

        wb.save(file_path)

    def _generate_summary(self, results: List[EvaluationResult]) -> Dict[str, Any]:
        """
        Generate summary statistics from evaluation results.

        Args:
            results: List of evaluation results

        Returns:
            Dictionary containing summary statistics
        """
        if not results:
            return {"total_files": 0}

        total_files = len(results)
        passed_files = sum(1 for r in results if r.passed)
        failed_files = total_files - passed_files
        skipped_files = sum(1 for r in results if r.metrics.get('comparison_skipped', False))
        files_with_references = sum(1 for r in results if r.metrics.get('reference_found', False))

        scores = [r.score for r in results if not r.metrics.get('comparison_skipped', False)]
        avg_score = sum(scores) / len(scores) if scores else 0

        content_total_comparisons = 0
        content_exact_matches = 0
        content_practical_matches = 0
        content_reordered_elements = 0
        content_tp = content_fp = content_fn = 0
        content_macro_precisions = []
        content_macro_recalls = []
        content_macro_f1s = []

        for result in results:
            summary_stats = result.metrics.get('summary_stats', {})

            content = summary_stats.get('content', {})
            content_total_comparisons += content.get('total_elements', 0)
            content_exact_matches += content.get('exact_matches', 0)
            content_practical_matches += content.get('practical_matches', 0)
            content_reordered_elements += content.get('reordered_elements', 0)
            content_metrics = content.get('metrics', summary_stats.get('content_metrics', {}))
            counts = content_metrics.get('counts', {})
            content_tp += counts.get('tp', 0)
            content_fp += counts.get('fp', 0)
            content_fn += counts.get('fn', 0)
            if content_metrics.get('available'):
                content_macro_precisions.append(content_metrics.get('macro_precision', 0))
                content_macro_recalls.append(content_metrics.get('macro_recall', 0))
                content_macro_f1s.append(content_metrics.get('macro_f1', 0))

        content_stats = {
            'total_comparisons': content_total_comparisons,
            'exact_matches': content_exact_matches,
            'practical_matches': content_practical_matches,
            'reordered_elements': content_reordered_elements,
            'exact_match_rate': (content_exact_matches / content_total_comparisons * 100) if content_total_comparisons > 0 else 0,
            'practical_match_rate': (content_practical_matches / content_total_comparisons * 100) if content_total_comparisons > 0 else 0,
            'files_with_references': files_with_references
        }

        if content_tp == 0 and content_fp == 0 and content_fn == 0:
            if content_total_comparisons == 0:
                content_precision = content_recall = content_f1 = 0.0
            else:
                content_precision = content_recall = content_f1 = 0.0
        else:
            content_precision = (content_tp / (content_tp + content_fp) * 100) if (content_tp + content_fp) > 0 else 0.0
            content_recall = (content_tp / (content_tp + content_fn) * 100) if (content_tp + content_fn) > 0 else 0.0
            if content_precision + content_recall > 0:
                content_f1 = (2 * content_precision * content_recall) / (content_precision + content_recall)
            else:
                content_f1 = 0.0

        if content_macro_precisions:
            macro_precision = sum(content_macro_precisions) / len(content_macro_precisions)
            macro_recall = sum(content_macro_recalls) / len(content_macro_recalls)
            macro_f1 = sum(content_macro_f1s) / len(content_macro_f1s)
        else:
            macro_precision = macro_recall = macro_f1 = 0.0

        content_stats.update({
            'precision': content_precision,
            'recall': content_recall,
            'f1': content_f1,
            'micro_precision': content_precision,
            'micro_recall': content_recall,
            'micro_f1': content_f1,
            'macro_precision': macro_precision,
            'macro_recall': macro_recall,
            'macro_f1': macro_f1
        })

        return {
            "total_files": total_files,
            "perfect_file_matches": passed_files,
            "failed_files": failed_files,
            "skipped_files": skipped_files,
            "pass_rate": (passed_files / total_files * 100) if total_files > 0 else 0,
            "average_score": avg_score,
            "min_score": min(scores) if scores else 0,
            "max_score": max(scores) if scores else 0,
            "content_stats": content_stats,
            "detailed_results": [
                {
                    "file_name": result.metrics.get('file_name', f"File_{i+1}"),
                    "file_path": result.metrics.get('file_path', ''),
                    "passed": result.passed,
                    "comparison_skipped": result.metrics.get('comparison_skipped', False),
                    "score": result.score,
                    "error_count": len(result.errors)
                }
                for i, result in enumerate(results)
            ]
        }

    def _add_element_issues_sheet(self, wb, results: List[EvaluationResult],
                                   header_font, header_fill, header_alignment):
        """Add element-level issues sheet (only non-exact matches)"""
        ws_issues = wb.create_sheet(title="Content Issues")

        issue_headers = [
            'File Name', 'Element Type', 'Element Index', 'Presence',
            'Match Type', 'Exact Match', 'Practical Match',
            'Content Category', 'Inclusion Ratio', 'Coverage Ratio',
            'LLM Content (Preview)', 'Reference Content (Preview)'
        ]
        for col, header in enumerate(issue_headers, 1):
            cell = ws_issues.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        current_row = 2
        for result in results:
            file_name = result.metrics.get('file_name', '')
            comparison_results = result.metrics.get('comparison_results', [])
            if not comparison_results:
                continue
            for element_comparison in comparison_results:
                element_type = element_comparison.get('element_type', 'unknown')
                content_comparisons = element_comparison.get('content_comparisons', [])
                for content_comp in content_comparisons:
                    element_index = content_comp.get('element_index', 0)
                    element_present = content_comp.get('element_present', 'unknown')
                    match_type = content_comp.get('match_type', 'unknown')
                    exact_match = content_comp.get('exact_match', False)
                    practical_match = content_comp.get('practical_match', False)
                    inclusion_ratio = content_comp.get('inclusion_ratio')
                    coverage_ratio = content_comp.get('coverage_ratio')
                    llm_content = content_comp.get('llm_content', '')
                    ref_content = content_comp.get('ref_content', '')
                    if exact_match:
                        continue
                    if element_present == 'both':
                        if exact_match:
                            content_category = 'Exact Match'
                        elif match_type == 'over_inclusion':
                            content_category = 'Over-inclusion (Practical)' if practical_match else 'Over-inclusion (Failed)'
                        elif match_type == 'under_inclusion':
                            content_category = 'Under-inclusion (Practical)' if practical_match else 'Under-inclusion (Failed)'
                        elif match_type == 'no_match':
                            content_category = 'No Match'
                        else:
                            content_category = 'Unknown'
                    elif element_present == 'llm_only':
                        content_category = 'Not Comparable (Extra in LLM)'
                    elif element_present == 'ref_only':
                        content_category = 'Not Comparable (Missing in LLM)'
                    else:
                        content_category = 'Unknown'
                    inclusion_str = f"{inclusion_ratio:.1%}" if inclusion_ratio is not None else "N/A"
                    coverage_str = f"{coverage_ratio:.1%}" if coverage_ratio is not None else "N/A"
                    llm_preview = self._format_content_preview(llm_content, content_comp.get('llm_element_info'))
                    ref_preview = ref_content[:100] + ('...' if len(ref_content) > 100 else '')
                    row_data = [
                        file_name,
                        element_type,
                        element_index + 1,
                        element_present,
                        match_type,
                        'Yes' if exact_match else 'No',
                        'Yes' if practical_match else 'No',
                        content_category,
                        inclusion_str,
                        coverage_str,
                        llm_preview,
                        ref_preview
                    ]
                    for col, value in enumerate(row_data, 1):
                        ws_issues.cell(row=current_row, column=col, value=value)
                    current_row += 1

        for column in ws_issues.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass  # Cell value cannot be converted to string
            adjusted_width = min(max_length + 2, 60)
            ws_issues.column_dimensions[column_letter].width = adjusted_width

    def _format_content_preview(self, content: str, element_info: dict = None) -> str:
        if not content or not content.strip():
            if element_info and element_info.get('has_children'):
                child_names = element_info.get('child_names', [])
                if child_names:
                    return f"contains {', '.join(child_names)} [child node(s)]"
            return "(empty)"
        return content[:100] + ('...' if len(content) > 100 else '')

    def _generate_error_breakdown(self, errors):
        breakdown = {}
        for error in errors:
            error_type = error.type.value if hasattr(error.type, 'value') else str(error.type)
            breakdown[error_type] = breakdown.get(error_type, 0) + 1
        return breakdown

    def _categorize_content_error(self, error) -> str:
        error_message = error.message.lower()
        if 'over-inclusion' in error_message:
            return "over_inclusion"
        elif 'under-inclusion' in error_message:
            return "under_inclusion"
        elif 'content mismatch' in error_message:
            return "content_mismatch"
        elif 'missing' in error_message:
            return "missing_element"
        else:
            return "other"

    def _categorize_errors(self, errors) -> Dict[str, int]:
        categories = {
            'content_errors': 0,
            'missing_elements': 0,
            'over_inclusion_occurrences': 0,
            'under_inclusion_occurrences': 0
        }
        for error in errors:
            category = self._categorize_content_error(error)
            if category == 'content_mismatch':
                categories['content_errors'] += 1
            elif category == 'missing_element':
                categories['missing_elements'] += 1
            elif category == 'over_inclusion':
                categories['over_inclusion_occurrences'] += 1
            elif category == 'under_inclusion':
                categories['under_inclusion_occurrences'] += 1
        return categories

    def _count_inclusion_occurrences_from_metrics(self, metrics: Dict[str, Any]) -> Dict[str, int]:
        """
        Count over/under-inclusion occurrences using comparison_results embedded in metrics.
        This is more reliable than parsing errors because practical matches are not always surfaced as errors.
        """
        counts = {
            'over_inclusion_occurrences': 0,
            'under_inclusion_occurrences': 0
        }
        comparison_results = metrics.get('comparison_results', [])
        if not isinstance(comparison_results, list):
            return counts

        for element_comparison in comparison_results:
            content_comparisons = element_comparison.get('content_comparisons', [])
            for comp in content_comparisons:
                match_type = comp.get('match_type')
                if match_type == 'over_inclusion':
                    counts['over_inclusion_occurrences'] += 1
                elif match_type == 'under_inclusion':
                    counts['under_inclusion_occurrences'] += 1
        return counts

    def _format_element_analysis(self, comparison_results: List[Dict[str, Any]]) -> Dict[str, Any]:
        element_analysis = {}
        for element_comp in comparison_results:
            element_type = element_comp.get('element_type', 'unknown')
            content_comparisons = element_comp.get('content_comparisons', [])
            element_summary = {
                'llm_count': element_comp.get('llm_count', 0),
                'ref_count': element_comp.get('ref_count', 0),
                'count_match': element_comp.get('count_match', False),
                'comparisons': []
            }
            for content_comp in content_comparisons:
                comparison_summary = {
                    'element_index': content_comp.get('element_index', 0),
                    'element_present': content_comp.get('element_present', 'unknown'),
                    'match_type': content_comp.get('match_type', 'unknown'),
                    'exact_match': content_comp.get('exact_match', False),
                    'fuzzy_match': content_comp.get('fuzzy_match', False),
                    'practical_match': content_comp.get('practical_match', False),
                    'inclusion_ratio': content_comp.get('inclusion_ratio'),
                    'coverage_ratio': content_comp.get('coverage_ratio'),
                    'analysis': content_comp.get('analysis', ''),
                    'llm_content_preview': content_comp.get('llm_content', '')[:100],
                    'ref_content_preview': content_comp.get('ref_content', '')[:100]
                }
                element_summary['comparisons'].append(comparison_summary)
            element_analysis[element_type] = element_summary
        return element_analysis

