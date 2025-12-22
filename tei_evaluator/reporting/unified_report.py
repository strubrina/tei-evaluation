# reporting/unified_report.py

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class UnifiedReporter:
    """Unified reporter that combines key metrics from all TEI evaluation dimensions"""

    def __init__(self, schema_mode: str = "both"):
        """
        Initialize unified reporter

        Args:
            schema_mode: Schema validation mode for overall scoring
                - "none": No schema validation impact on score
                - "tei": TEI schema validation only
                - "project": Project schema validation only
                - "both": Both TEI and project schema validation (default)
        """
        self.dimension0_results = {}
        self.dimension1_results = {}
        self.dimension2_results = {}
        self.dimension3_results = {}
        self.dimension4_results = {}
        self.schema_mode = schema_mode

    def load_dimension0_results(self, results_dir: str):
        """Load Dimension 0 (wellformedness) results"""
        results_path = Path(results_dir)

        wellformedness_file = results_path / "json" / "d0_wellformedness_report.json"
        if wellformedness_file.exists():
            print(f"[INFO] Loading Dimension 0 wellformedness from: {wellformedness_file.name}")
            self._extract_dimension0_data(wellformedness_file)

    def load_dimension1_results(self, results_dir: str):
        """Load Dimension 1 (source fidelity) results"""
        results_path = Path(results_dir)

        # Prefer new location under json/
        source_fidelity_file = results_path / "json" / "d1_source_fidelity_report.json"
        if not source_fidelity_file.exists():
            # Fallback to legacy location at root
            legacy_path = results_path / "d1_source_fidelity_report.json"
            source_fidelity_file = legacy_path if legacy_path.exists() else source_fidelity_file

        if source_fidelity_file.exists():
            print(f"[INFO] Loading Dimension 1 source fidelity from: {source_fidelity_file.name}")
            self._extract_dimension1_data(source_fidelity_file)

    def load_dimension2_results(self, results_dir: str):
        """Load Dimension 2 (schema validation) results"""
        results_path = Path(results_dir)
        # Prefer file based on schema_mode to avoid picking stale combined reports
        tei_only = results_path / "json" / "d2_tei_validation_report.json"
        project_only = results_path / "json" / "d2_project_validation_report.json"
        combined = results_path / "json" / "d2_schema_validation_report.json"

        if self.schema_mode == "tei":
            preferred_order = [tei_only, combined, project_only]
        elif self.schema_mode == "project":
            preferred_order = [project_only, combined, tei_only]
        else:  # "both" or any other value defaults to combined
            preferred_order = [combined, project_only, tei_only]

        detailed_file = None
        for candidate in preferred_order:
            if candidate.exists():
                detailed_file = candidate
                break

        if detailed_file.exists():
            self._extract_dimension2_data(detailed_file)

    def load_dimension3_results(self, results_dir: str):
        """Load Dimension 3 (structural comparison) results"""
        results_path = Path(results_dir)
        # Updated JSON filename produced by d3_report.py
        detailed_file = results_path / "json" / "d3_detailed_report.json"
        if not detailed_file.exists():
            # Fallbacks for legacy names/locations
            legacy_json = results_path / "json" / "d3_structural_fidelity_report.json"
            legacy_root = results_path / "d3_detailed_report.json"
            detailed_file = legacy_json if legacy_json.exists() else (legacy_root if legacy_root.exists() else detailed_file)

        if detailed_file.exists():
            self._extract_dimension3_data(detailed_file)

    def load_dimension4_results(self, results_dir: str):
        """Load Dimension 4 (content matching) results"""
        results_path = Path(results_dir)
        # Updated JSON filename produced by d4_report.py
        detailed_file = results_path / "json" / "d4_detailed_report.json"
        if not detailed_file.exists():
            # Fallbacks for legacy names/locations
            legacy_json = results_path / "json" / "d4_semantic_validation_report.json"
            legacy_root = results_path / "d4_detailed_report.json"
            detailed_file = legacy_json if legacy_json.exists() else (legacy_root if legacy_root.exists() else detailed_file)

        if detailed_file.exists():
            self._extract_dimension4_data(detailed_file)

    def _extract_dimension0_data(self, file_path: Path):
        """Extract key wellformedness metrics from Dimension 0 data"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "files" not in data:
            return

        for file_data in data["files"]:
            if isinstance(file_data, str):  # Skip section headers
                continue

            file_name = file_data.get("file_name", "")
            if not file_name:
                continue

            if file_name not in self.dimension0_results:
                self.dimension0_results[file_name] = {}

            # Read from flattened JSON structure
            self.dimension0_results[file_name].update({
                "well_formed": file_data.get("well_formed", False),
                "score": file_data.get("xml_score", 0),
                "xml_errors": file_data.get("xml_errors", 0),
                "file_not_found": file_data.get("file_not_found", False),
                "non_xml_errors": len(file_data.get("non_xml_errors", [])) if isinstance(file_data.get("non_xml_errors", []), list) else 0
            })

    def _extract_dimension1_data(self, file_path: Path):
        """Extract key source fidelity metrics from Dimension 1 data"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "files" not in data:
            return

        for file_data in data["files"]:
            if isinstance(file_data, str):  # Skip section headers
                continue

            file_name = file_data.get("file_name", "")
            if not file_name:
                continue

            if file_name not in self.dimension1_results:
                self.dimension1_results[file_name] = {}

            # Read from flattened JSON structure
            self.dimension1_results[file_name].update({
                "content_preserved": file_data.get("content_preserved", False),
                "content_score": file_data.get("content_score", 0),
                "content_errors": file_data.get("differences_count", 0)
            })

    def _extract_dimension2_data(self, file_path: Path):
        """Extract key schema validation metrics from Dimension 2 data"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "files" not in data:
            return

        for file_data in data["files"]:
            # Skip string headers (e.g., "COMBINED (TEI + PROJECT):")
            if isinstance(file_data, str):
                continue

            file_name = file_data.get("file_name", "")
            if not file_name:
                continue

            self.dimension2_results[file_name] = {}

            # Read from flattened JSON structure (updated)
            self.dimension2_results[file_name].update({
                "tei_valid": file_data.get("tei_valid", "N/A"),
                "tei_score": file_data.get("tei_score", "N/A"),
                "tei_validity_reference": file_data.get("tei_validity_reference", "N/A"),
                "tei_validity_match": file_data.get("tei_validity_match", "N/A"),
                "project_valid": file_data.get("project_valid", "N/A"),
                "project_score": file_data.get("project_score", "N/A"),
                "project_validity_reference": file_data.get("project_validity_reference", "N/A"),
                "project_validity_match": file_data.get("project_validity_match", "N/A")
            })

    def _extract_dimension3_data(self, file_path: Path):
        """Extract key structural comparison metrics from Dimension 3 data"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "files" not in data:
            return

        for file_data in data["files"]:
            # Skip string headers
            if isinstance(file_data, str):
                continue

            file_name = file_data.get("file_name", "")
            if not file_name:
                continue

            self.dimension3_results[file_name] = {}

            # Completeness and structural data
            self.dimension3_results[file_name].update({
                "completeness_match": file_data.get("completeness_match", False),
                "completeness_score": file_data.get("completeness_score", 0),
                "structural_match": file_data.get("structural_match", False),
                "structural_score": file_data.get("structural_score", 0),
                "tree_edit_distance": file_data.get("tree_edit_distance"),
                "tree_edit_normalized": file_data.get("tree_edit_normalized"),
                "tree_similarity": file_data.get("ted_similarity", file_data.get("tree_similarity")),
                "tree_lcs_similarity": file_data.get("tree_lcs_similarity", file_data.get("structural_score", 0))
            })

    def _extract_dimension4_data(self, file_path: Path):
        """Extract key content matching metrics from Dimension 4 data"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        if "files" not in data:
            return

        for file_data in data["files"]:
            # Skip string headers
            if isinstance(file_data, str):
                continue

            file_name = file_data.get("file_name", "")
            if not file_name:
                continue

            self.dimension4_results[file_name] = {}

            # Read from flattened JSON structure (updated)
            self.dimension4_results[file_name].update({
                "content_match": file_data.get("semantic_match", False),
                "content_score": file_data.get("score", 0),
                "exact_content_matches": file_data.get("exact_matches", 0),
                "practical_content_matches": file_data.get("practical_matches", 0),
                "total_elements": file_data.get("compared_elements", 0),
                "content_precision": file_data.get("content_precision", 0),
                "content_recall": file_data.get("content_recall", 0),
                "content_f1": file_data.get("content_f1", 0),
                "content_macro_f1": file_data.get("content_macro_f1", 0),
                "content_micro_f1": file_data.get("content_micro_f1", 0)
            })

    def _calculate_d4_context_aware_score(self, d1_content_preserved: bool,
                                          exact_matches: int, practical_matches: int,
                                          total_elements: int, original_score: float) -> float:
        """
        Calculate context-aware D4 score based on D1 content preservation status.

        This recalculates the D4 score with adjusted weights for practical matches:
        - If D1 content is preserved: practical matches likely represent boundary issues → weight 0.8
        - If D1 content has issues: practical matches less trustworthy → weight 0.6

        IMPORTANT: This is ONLY used for the unified overall score calculation.
        Individual D4 reports retain their original scoring with fixed weights (practical_weight=0.6)
        to ensure consistency and comparability across evaluations.

        Design Rationale:
        - Individual dimension reports should be stable and independent
        - Unified score can incorporate cross-dimensional insights
        - Researchers can compare D4 individual scores across files/models
        - Overall score reflects contextual quality interpretation

        For the original D4 scoring logic, see:
        tei_evaluator/core/dimension4_semantic.py::calculate_score()

        Args:
            d1_content_preserved: Whether D1 content preservation passed
            exact_matches: Number of exact content matches
            practical_matches: Number of practical content matches
            total_elements: Total number of elements compared
            original_score: Original D4 score (for fallback)

        Returns:
            Context-aware adjusted D4 score (only for unified overall score)
        """
        # If we don't have the data, return original score
        if total_elements == 0:
            return original_score

        # Determine practical match weight based on D1 content preservation
        if d1_content_preserved:
            practical_weight = 0.8  # Higher weight - likely boundary issues
        else:
            practical_weight = 0.6  # Lower weight - possible hallucinations

        # Recalculate weighted matches
        exact_weight = 1.0
        weighted_matches = (exact_matches * exact_weight) + (practical_matches * practical_weight)

        # Calculate match ratio and score
        match_ratio = weighted_matches / (total_elements * exact_weight)
        adjusted_score = match_ratio * 100

        # Cap at 100
        return min(100.0, adjusted_score)

    def _calculate_schema_factor(self, tei_valid: bool, project_valid: bool) -> float:
        """
        Calculate schema compliance factor based on validation mode.

        Schema validation is treated as a quality gate/requirement, not a bonus.
        - Valid schemas: factor 1.0 (meets requirements - no bonus)
        - Invalid schemas: factor 0.75 (25% penalty for non-compliance)

        Schema violations receive heavy penalties because they are harder to
        automatically correct and require manual intervention (e.g., wrong element
        choices, incorrect nesting structures).

        Args:
            tei_valid: Whether TEI schema validation passed
            project_valid: Whether project schema validation passed

        Returns:
            Multiplication factor (0.75 or 1.0)
        """
        if self.schema_mode == "none":
            return 1.0

        elif self.schema_mode == "tei":
            return 1.0 if tei_valid else 0.75

        elif self.schema_mode == "project":
            return 1.0 if project_valid else 0.75

        elif self.schema_mode == "both":
            # Binary: either meets all requirements or fails compliance
            if tei_valid and project_valid:
                return 1.0  # Full compliance (baseline)
            else:
                return 0.75  # 25% penalty for non-compliance

        return 1.0  # Default

    def _calculate_overall_score(self, well_formed: bool, d1_content_score: float,
                                 d1_content_preserved: bool, d3_lcs: float,
                                 d4_content_score: float,
                                 d4_exact_matches: int, d4_practical_matches: int,
                                 d4_total_elements: int,
                                 d4_macro_f1: float,
                                 tei_valid: bool, project_valid: bool) -> Dict[str, Any]:
        """
        Calculate new overall score using tiered evaluation system.

        Tier 1: Core Quality Dimensions (weighted equally)
        - D1 Content Preservation: 33%
        - D3 LCS Structural: 33% (LCS-based structural similarity)
        - D4 Content Matching: 33%

        Tier 2: Schema Compliance Adjustment (0.90 - 1.0x multiplier)
        - Schema validation treated as quality gate (penalties only, no bonuses)
        - Valid schemas: 1.0x (meets baseline requirements)
        - Invalid schemas: 0.90-0.95x (penalty for not meeting standards)

        Returns:
            Dictionary with overall_score and scoring breakdown
        """
        # Tier 0: Validity Gate
        if not well_formed:
            return {
                "overall_score": None,
                "status": "INVALID",
                "core_quality_score": None,
                "d4_adjusted_score_OLD": None,
                "d4_macro_f1": None,
                "schema_factor": None,
                "breakdown": "File is not well-formed XML and cannot be processed"
            }

        # Tier 1: Core Quality Dimensions (weighted independently)

        # D4 F1-based Score (new Macro F1 approach)
        # Use the Macro F1 score directly (already in 0-100 range)
        d4_f1_score = d4_macro_f1 if d4_macro_f1 > 0 else d4_content_score  # Fallback to content_score if F1 not available

        # Core Quality Score (using LCS for structural, no completeness)
        # Equal weights: 33.33% each
        core_quality_score = (
            d1_content_score * (1/3) +     # Content preservation: 33%
            d3_lcs * (1/3) +               # LCS structural similarity: 33%
            d4_f1_score * (1/3)            # Content matching (F1-based): 33%
        )

        # Tier 2: Schema Compliance Adjustment
        schema_factor = self._calculate_schema_factor(tei_valid, project_valid)

        # Calculate final overall score (no longer needs capping since schema_factor ≤ 1.0)
        overall_score = core_quality_score * schema_factor

        return {
            "overall_score": round(overall_score, 2),
            "status": "VALID",
            "core_quality_score": round(core_quality_score, 2),
            "d4_adjusted_semantic_score": round(d4_f1_score, 2),  # Now using F1-based score
            "schema_factor": schema_factor,
            "breakdown": {
                "d1_weight": round(1/3, 4),
                "d3_lcs_weight": round(1/3, 4),
                "d4_weight": round(1/3, 4),
                "schema_mode": self.schema_mode
            }
        }

    def generate_unified_report(self, output_dir: str):
        """Generate unified Excel and JSON reports combining all dimensions.

        Excel is saved at the processing root (output_dir).
        JSON is saved under output_dir/json/.
        """
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        json_dir = output_path / "json"
        json_dir.mkdir(exist_ok=True)

        # Get all unique file names across all dimensions
        all_files = set()
        all_files.update(self.dimension0_results.keys())
        all_files.update(self.dimension1_results.keys())
        all_files.update(self.dimension2_results.keys())
        all_files.update(self.dimension3_results.keys())
        all_files.update(self.dimension4_results.keys())

        if not all_files:
            print("[WARNING] No evaluation results found to combine")
            return

        # Generate unified data
        unified_data = []
        for file_name in sorted(all_files):
            file_data = self._combine_file_data(file_name)
            unified_data.append(file_data)

        # Generate summary statistics
        summary_stats = self._generate_summary_stats(unified_data)

        # Save Excel report with summary (root)
        self._save_unified_excel(unified_data, summary_stats, output_path / "unified_evaluation_summary.xlsx")

        # Save JSON report (json/)
        self._save_unified_json(unified_data, json_dir / "unified_evaluation_summary.json")

        # Create heatmap visualization
        try:
            from .visualization import TEIEvaluationVisualizer
            visualizer = TEIEvaluationVisualizer(str(output_path))
            heatmap_path = visualizer.create_unified_evaluation_heatmap(unified_data)
            print(f"[OUTPUT]    - {heatmap_path.name}")
        except Exception as e:
            print(f"[WARNING] Could not generate heatmap: {e}")

        # Create visualizations (disabled - use generate_visualizations.py instead)
        # self._create_visualizations(unified_data, summary_stats, output_path)

        print(f"[OUTPUT] Unified reports saved to: {output_path}")
        print(f"[OUTPUT]    - unified_evaluation_summary.xlsx")
        print(f"[OUTPUT]    - json/unified_evaluation_summary.json")
        print(f"[INFO] To generate additional visualizations, run: python generate_visualizations.py")

    def _combine_file_data(self, file_name: str) -> Dict[str, Any]:
        """Combine data for a single file from all evaluation dimensions"""
        dimension0_data = self.dimension0_results.get(file_name, {})
        dimension1_data = self.dimension1_results.get(file_name, {})
        dimension2_data = self.dimension2_results.get(file_name, {})
        dimension3_data = self.dimension3_results.get(file_name, {})
        dimension4_data = self.dimension4_results.get(file_name, {})

        # Get values with defaults
        well_formed = dimension0_data.get("well_formed", False)
        content_preserved = dimension1_data.get("content_preserved", False)
        tei_valid = dimension2_data.get("tei_valid", "N/A")
        project_valid = dimension2_data.get("project_valid", "N/A")

        # Calculate scores
        d0_score = dimension0_data.get("score", 0)
        d1_content_score = dimension1_data.get("content_score", 0)
        d2_tei_score = dimension2_data.get("tei_score", "N/A")
        d2_project_score = dimension2_data.get("project_score", "N/A")
        d3_completeness = dimension3_data.get("completeness_score", 0)
        d3_structural = dimension3_data.get("structural_score", 0)
        d3_lcs = dimension3_data.get("tree_lcs_similarity", 0)
        # Convert LCS from 0-1 range to 0-100 range if needed
        if isinstance(d3_lcs, (int, float)) and d3_lcs <= 1.0 and d3_lcs >= 0:
            d3_lcs = d3_lcs * 100
        d4_content = dimension4_data.get("content_score", 0)
        d4_exact_matches = dimension4_data.get("exact_content_matches", 0)
        d4_practical_matches = dimension4_data.get("practical_content_matches", 0)
        d4_total_elements = dimension4_data.get("total_elements", 0)
        d4_macro_f1 = dimension4_data.get("content_macro_f1", 0)

        # NEW TIERED SCORING (using LCS instead of structural_score)
        scoring_result = self._calculate_overall_score(
            well_formed=well_formed,
            d1_content_score=d1_content_score if isinstance(d1_content_score, (int, float)) else 0,
            d1_content_preserved=content_preserved,
            d3_lcs=d3_lcs if isinstance(d3_lcs, (int, float)) else 0,
            d4_content_score=d4_content,
            d4_exact_matches=d4_exact_matches if isinstance(d4_exact_matches, int) else 0,
            d4_practical_matches=d4_practical_matches if isinstance(d4_practical_matches, int) else 0,
            d4_total_elements=d4_total_elements if isinstance(d4_total_elements, int) else 0,
            d4_macro_f1=d4_macro_f1 if isinstance(d4_macro_f1, (int, float)) else 0,
            tei_valid=tei_valid if isinstance(tei_valid, bool) else False,
            project_valid=project_valid if isinstance(project_valid, bool) else False
        )

        overall_score_valid = scoring_result.get("overall_score")
        overall_score_all_files = (
            overall_score_valid if scoring_result.get("status") == "VALID" and isinstance(overall_score_valid, (int, float))
            else 0.0
        )

        return {
            # File identification
            "file_name": file_name,
            # file_id removed from detailed results

            # Dimension 0 - Wellformedness
            "d0_file_not_found": dimension0_data.get("file_not_found", False),
            "d0_well_formed": well_formed,
            "d0_score": d0_score,
            "d0_xml_errors": dimension0_data.get("xml_errors", "N/A"),
            "d0_non_xml_errors": dimension0_data.get("non_xml_errors", 0),

            # Dimension 1 - Source Fidelity
            "d1_source_fidelity": content_preserved,
            "d1_source_fidelity_score": d1_content_score,
            "d1_content_errors": dimension1_data.get("content_errors", "N/A"),

            # Dimension 2 - Schema Validation
            "d2_tei_valid": tei_valid if isinstance(tei_valid, bool) else "N/A",
            "d2_tei_score": d2_tei_score if isinstance(d2_tei_score, (int, float)) else "N/A",
            "d2_tei_validity_reference": dimension2_data.get("tei_validity_reference", "N/A"),
            "d2_tei_validity_match": dimension2_data.get("tei_validity_match", "N/A"),
            "d2_project_valid": project_valid if isinstance(project_valid, bool) else "N/A",
            "d2_project_score": d2_project_score if isinstance(d2_project_score, (int, float)) else "N/A",
            "d2_project_validity_reference": dimension2_data.get("project_validity_reference", "N/A"),
            "d2_project_validity_match": dimension2_data.get("project_validity_match", "N/A"),

            # Dimension 3 - Structural Comparison
            "d3_structural_match": dimension3_data.get("structural_match", "N/A"),
            "d3_structural_score": d3_structural,
            # d3_lcs_score removed from detailed results

            # Dimension 4 - Content Matching
            "d4_content_match": dimension4_data.get("content_match", "N/A"),
            "d4_semantic_score": d4_content,  # Renamed from d4_content_score
            "d4_exact_content_matches": d4_exact_matches,
            "d4_practical_content_matches": d4_practical_matches,
            "d4_total_elements": d4_total_elements,
            "d4_adjusted_semantic_score": scoring_result.get("d4_adjusted_semantic_score", "N/A"),

            # Overall Scores and Metadata
            "core_quality_score": scoring_result.get("core_quality_score", "N/A"),
            "overall_score": scoring_result.get("overall_score"),
            "overall_score_all_files": overall_score_all_files,
            "evaluation_status": scoring_result.get("status", "UNKNOWN"),  # VALID or INVALID
            "schema_factor": scoring_result.get("schema_factor", "N/A"),
            "schema_mode": self.schema_mode,
        }


    def _save_unified_excel(self, data: List[Dict[str, Any]], summary_stats: Dict[str, Any], file_path: Path):
        """Save unified data as Excel with Summary and Detailed Results sheets"""
        if not data:
            return

        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not available
            csv_path = file_path.with_suffix('.csv')
            self._save_unified_csv(data, csv_path)
            return

        wb = Workbook()

        # ========== SHEET 1: SUMMARY ==========
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Define styles
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        section_font = Font(bold=True, size=12, color="FFFFFF")
        section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        label_font = Font(bold=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        row = 1

        # Title
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="UNIFIED TEI EVALUATION SUMMARY")
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 2

        # Overall statistics
        ws_summary.cell(row=row, column=1, value="Total Files Evaluated:").font = label_font
        ws_summary.cell(row=row, column=2, value=summary_stats.get("total_files", 0))
        row += 1

        # Display selected schema mode prominently under totals
        overall_meta = summary_stats.get("overall_tiered_scoring", {})
        schema_mode_value = overall_meta.get("schema_mode", self.schema_mode)
        # Human-readable label
        mode_map = {
            "none": "None",
            "tei": "TEI Only",
            "project": "Project Only",
            "both": "Combined (TEI + Project)"
        }
        ws_summary.cell(row=row, column=1, value="Schema Mode:").font = label_font
        ws_summary.cell(row=row, column=2, value=mode_map.get(schema_mode_value, schema_mode_value))
        row += 2

        # ===== DIMENSION 0: WELLFORMEDNESS =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="DIMENSION 0: WELLFORMEDNESS")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        d1_wf = summary_stats.get("dimension1_wellformedness", {})
        ws_summary.cell(row=row, column=1, value="Total Files:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_wf.get("total_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Passed:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_wf.get("passed_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Failed:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_wf.get("failed_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Pass Rate:").font = label_font
        pass_rate_cell = ws_summary.cell(row=row, column=2, value=f"{d1_wf.get('pass_rate', 0):.2f}%")
        pass_rate_val = d1_wf.get('pass_rate', 0)
        if pass_rate_val >= 90:
            pass_rate_cell.fill = green_fill
        elif pass_rate_val >= 70:
            pass_rate_cell.fill = yellow_fill
        else:
            pass_rate_cell.fill = red_fill
        row += 1

        ws_summary.cell(row=row, column=1, value="Total Errors:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_wf.get("total_errors", 0))
        row += 2

        # Additional non-XML visibility for D0
        ws_summary.cell(row=row, column=1, value="Files Not Found:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_wf.get("files_not_found", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Non-XML Errors:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_wf.get("non_xml_errors_total", 0))
        row += 2

        # ===== DIMENSION 1: CONTENT PRESERVATION =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="DIMENSION 1: SOURCE FIDELITY")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        d1_cp = summary_stats.get("dimension1_content", {})
        ws_summary.cell(row=row, column=1, value="Total Files:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_cp.get("total_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Passed:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_cp.get("passed_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Failed:").font = label_font
        ws_summary.cell(row=row, column=2, value=d1_cp.get("failed_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Pass Rate:").font = label_font
        pass_rate_cell = ws_summary.cell(row=row, column=2, value=f"{d1_cp.get('pass_rate', 0):.2f}%")
        pass_rate_val = d1_cp.get('pass_rate', 0)
        if pass_rate_val >= 90:
            pass_rate_cell.fill = green_fill
        elif pass_rate_val >= 70:
            pass_rate_cell.fill = yellow_fill
        else:
            pass_rate_cell.fill = red_fill
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d1_cp.get('average_content_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d1_cp.get('min_content_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d1_cp.get('max_content_score', 0):.2f}")
        row += 2

        # ===== DIMENSION 2: TEI VALIDATION =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="DIMENSION 2: TEI SCHEMA VALIDATION")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        d2_tei = summary_stats.get("dimension2_tei", {})
        if d2_tei.get("total_files", 0) > 0:
            ws_summary.cell(row=row, column=1, value="Total Files:").font = label_font
            ws_summary.cell(row=row, column=2, value=d2_tei.get("total_files", 0))
            row += 1

            ws_summary.cell(row=row, column=1, value="Valid:").font = label_font
            ws_summary.cell(row=row, column=2, value=d2_tei.get("valid_files", 0))
            row += 1

            ws_summary.cell(row=row, column=1, value="Invalid:").font = label_font
            ws_summary.cell(row=row, column=2, value=d2_tei.get("invalid_files", 0))
            row += 1

            ws_summary.cell(row=row, column=1, value="Pass Rate:").font = label_font
            pass_rate_cell = ws_summary.cell(row=row, column=2, value=f"{d2_tei.get('pass_rate', 0):.2f}%")
            pass_rate_val = d2_tei.get('pass_rate', 0)
            if pass_rate_val >= 90:
                pass_rate_cell.fill = green_fill
            elif pass_rate_val >= 70:
                pass_rate_cell.fill = yellow_fill
            else:
                pass_rate_cell.fill = red_fill
            row += 2
        else:
            ws_summary.cell(row=row, column=1, value="No TEI validation (N/A)").font = label_font
            row += 2

        # ===== DIMENSION 2: PROJECT VALIDATION =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="DIMENSION 2: PROJECT SCHEMA VALIDATION")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        d2_proj = summary_stats.get("dimension2_project", {})
        if d2_proj.get("total_files", 0) > 0:
            ws_summary.cell(row=row, column=1, value="Total Files:").font = label_font
            ws_summary.cell(row=row, column=2, value=d2_proj.get("total_files", 0))
            row += 1

            ws_summary.cell(row=row, column=1, value="Valid:").font = label_font
            ws_summary.cell(row=row, column=2, value=d2_proj.get("valid_files", 0))
            row += 1

            ws_summary.cell(row=row, column=1, value="Invalid:").font = label_font
            ws_summary.cell(row=row, column=2, value=d2_proj.get("invalid_files", 0))
            row += 1

            ws_summary.cell(row=row, column=1, value="Pass Rate:").font = label_font
            pass_rate_cell = ws_summary.cell(row=row, column=2, value=f"{d2_proj.get('pass_rate', 0):.2f}%")
            pass_rate_val = d2_proj.get('pass_rate', 0)
            if pass_rate_val >= 90:
                pass_rate_cell.fill = green_fill
            elif pass_rate_val >= 70:
                pass_rate_cell.fill = yellow_fill
            else:
                pass_rate_cell.fill = red_fill
            row += 2
        else:
            ws_summary.cell(row=row, column=1, value="No Project validation (N/A)").font = label_font
            row += 2

        # ===== DIMENSION 3: STRUCTURAL COMPARISON =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="DIMENSION 3: STRUCTURAL COMPARISON")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        d3 = summary_stats.get("dimension3_structural", {})
        ws_summary.cell(row=row, column=1, value="Total Files:").font = label_font
        ws_summary.cell(row=row, column=2, value=d3.get("total_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Structural Matches:").font = label_font
        ws_summary.cell(row=row, column=2, value=d3.get("matched_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Structural Differences:").font = label_font
        ws_summary.cell(row=row, column=2, value=d3.get("different_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Structural Match Rate:").font = label_font
        pass_rate_cell = ws_summary.cell(row=row, column=2, value=f"{d3.get('match_rate', 0):.2f}%")
        pass_rate_val = d3.get('match_rate', 0)
        if pass_rate_val >= 90:
            pass_rate_cell.fill = green_fill
        elif pass_rate_val >= 70:
            pass_rate_cell.fill = yellow_fill
        else:
            pass_rate_cell.fill = red_fill
        row += 1

        ws_summary.cell(row=row, column=1, value="Avg Structural Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d3.get('average_structural_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Structural Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d3.get('min_structural_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Structural Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d3.get('max_structural_score', 0):.2f}")
        row += 1

        # Blank line separator
        row += 2

        # ===== DIMENSION 4: CONTENT MATCHING =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="DIMENSION 4: SEMANTIC CONTENT MATCHING")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        d4 = summary_stats.get("dimension4_content", {})
        ws_summary.cell(row=row, column=1, value="Total Files:").font = label_font
        ws_summary.cell(row=row, column=2, value=d4.get("total_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Matched:").font = label_font
        ws_summary.cell(row=row, column=2, value=d4.get("passed_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Different:").font = label_font
        ws_summary.cell(row=row, column=2, value=d4.get("failed_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Match Rate:").font = label_font
        pass_rate_cell = ws_summary.cell(row=row, column=2, value=f"{d4.get('pass_rate', 0):.2f}%")
        pass_rate_val = d4.get('pass_rate', 0)
        if pass_rate_val >= 90:
            pass_rate_cell.fill = green_fill
        elif pass_rate_val >= 70:
            pass_rate_cell.fill = yellow_fill
        else:
            pass_rate_cell.fill = red_fill
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d4.get('average_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d4.get('min_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Score:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d4.get('max_score', 0):.2f}")
        row += 1

        d4_content = d4.get("content_stats", {})
        ws_summary.cell(row=row, column=1, value="Total Element Exact Matches:").font = label_font
        ws_summary.cell(row=row, column=2, value=d4_content.get("total_exact_matches", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Total Element Practical Matches:").font = label_font
        ws_summary.cell(row=row, column=2, value=d4_content.get("total_practical_matches", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Exact Match Rate:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d4_content.get('exact_match_rate', 0):.2f}%")
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Exact Match Rate:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d4_content.get('exact_match_rate', 0):.2f}%")
        row += 1
        ws_summary.cell(row=row, column=1, value="Average Practical Match Rate:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{d4_content.get('practical_match_rate', 0):.2f}%")
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Relaxed Content Matching Score").font = label_font
        overall = summary_stats.get("overall_tiered_scoring", {})
        ws_summary.cell(row=row, column=2, value=f"{overall.get('average_d4_adjusted_semantic', 0):.2f}")
        row += 1
        row += 1

        # ===== OVERALL TIERED SCORING =====
        ws_summary.merge_cells(f'A{row}:C{row}')
        cell = ws_summary.cell(row=row, column=1, value="OVERALL TIERED SCORING")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")
        row += 1

        overall = summary_stats.get("overall_tiered_scoring", {})
        ws_summary.cell(row=row, column=1, value="Schema Mode:").font = label_font
        ws_summary.cell(row=row, column=2, value=overall.get("schema_mode", "N/A"))
        row += 1

        ws_summary.cell(row=row, column=1, value="Valid Files:").font = label_font
        ws_summary.cell(row=row, column=2, value=overall.get("valid_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Invalid Files:").font = label_font
        ws_summary.cell(row=row, column=2, value=overall.get("invalid_files", 0))
        row += 1

        ws_summary.cell(row=row, column=1, value="Validity Rate:").font = label_font
        validity_rate_cell = ws_summary.cell(row=row, column=2, value=f"{overall.get('validity_rate', 0):.2f}%")
        validity_rate_val = overall.get('validity_rate', 0)
        if validity_rate_val >= 90:
            validity_rate_cell.fill = green_fill
        elif validity_rate_val >= 70:
            validity_rate_cell.fill = yellow_fill
        else:
            validity_rate_cell.fill = red_fill
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Overall Score (only valid XML):").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('average_overall_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Overall Score (all files):").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('average_overall_score_all_files', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Overall Score (only valid XML):").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('min_overall_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Overall Score (all files):").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('min_overall_score_all_files', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Overall Score (only valid XML):").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('max_overall_score', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Overall Score (all files):").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('max_overall_score_all_files', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Core Quality Score").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('average_core_quality', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Core Quality Score").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('max_core_quality', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Core Quality Score").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('min_core_quality', 0):.2f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Average Schema Factor:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('average_schema_factor', 1.0):.3f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Min Schema Factor:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('min_schema_factor', 1.0):.3f}")
        row += 1

        ws_summary.cell(row=row, column=1, value="Max Schema Factor:").font = label_font
        ws_summary.cell(row=row, column=2, value=f"{overall.get('max_schema_factor', 1.0):.3f}")

        # Adjust column widths for summary sheet
        ws_summary.column_dimensions['A'].width = 35
        ws_summary.column_dimensions['B'].width = 20
        ws_summary.column_dimensions['C'].width = 15

        # ========== SHEET 2: DETAILED RESULTS ==========
        ws_details = wb.create_sheet("Detailed Results")

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Color coding styles for booleans and scores
        true_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
        false_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for scores
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for scores
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for scores

        # Write headers and build column maps
        # Columns to remove coloring from (by header name)
        no_coloring_headers = [
            'd0_file_not_found',  # Column B - remove coloring
            'd2_tei_score',  # Column K - remove coloring
            'd2_tei_validity_reference',  # Column L - remove coloring
            'd3_structural_score',  # Column S - remove coloring
        ]

        # Special column: Column G - needs special handling (green for 0, red for >0, not boolean)
        # This is likely d1_source_fidelity_score or another numeric column
        special_count_headers = [
            'd1_source_fidelity_score',  # Column G - special: green for 0, red for >0
        ]

        # Columns that need boolean coloring
        boolean_column_names = [
            'd0_well_formed',  # Column C - needs coloring (boolean)
            'd1_source_fidelity',  # Boolean column
            'd2_tei_valid', 'd2_tei_validity_match',
            'd2_project_valid',  # Column N - needs coloring (boolean)
            'd2_project_validity_reference', 'd2_project_validity_match',
            'd3_structural_match',
            'd4_content_match'
        ]

        score_column_names = [
            'd2_project_score',
            'd4_semantic_score', 'd4_adjusted_semantic_score',
            'overall_score', 'overall_score_all_files', 'core_quality_score'
        ]
        # Note: d1_source_fidelity_score is in special_count_headers, not score_column_names

        # Map column names to indices
        boolean_columns = {}  # header_name -> column_index
        score_columns = {}  # header_name -> column_index
        header_to_col = {}  # header_name -> column_index mapping

        if data:
            # Determine columns to include: drop columns where all values are "N/A"
            all_keys = list(data[0].keys())
            include_keys = []
            for key in all_keys:
                if any(row.get(key) != "N/A" for row in data):
                    include_keys.append(key)

            for col, header in enumerate(include_keys, 1):
                cell = ws_details.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

                header_to_col[header] = col

                # Track column types (exclude columns that should not be colored)
                if header not in no_coloring_headers:
                    if header in boolean_column_names:
                        boolean_columns[header] = col
                    elif header in score_column_names:
                        score_columns[header] = col

        # Write data rows
        for row_idx, row_data in enumerate(data, 2):
            col_idx = 1
            for key in include_keys:
                value = row_data.get(key)
                cell = ws_details.cell(row=row_idx, column=col_idx, value=value)

                # Apply color coding based on header name
                if key not in no_coloring_headers:
                    # Special handling for column G (count column: green for 0, red for >0)
                    if key in special_count_headers:
                        if isinstance(value, (int, float)) and value != "N/A":
                            if value == 0:
                                cell.fill = green_fill
                            elif value > 0:
                                cell.fill = red_fill
                        # N/A values remain uncolored
                    elif key in boolean_columns:
                        # Boolean columns: green for True, red for False
                        if value is True:
                            cell.fill = true_fill
                        elif value is False:
                            cell.fill = false_fill
                        # N/A or other values remain uncolored
                    elif key in score_columns:
                        # Score columns: color by value
                        if isinstance(value, (int, float)) and value != "N/A":
                            if value >= 90:
                                cell.fill = green_fill
                            elif value >= 70:
                                cell.fill = yellow_fill
                            else:
                                cell.fill = red_fill
                        # N/A or other values remain uncolored

                col_idx += 1

        # Auto-adjust column widths for detailed results
        for column in ws_details.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, AttributeError):
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_details.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def _create_visualizations(self, data: List[Dict[str, Any]],
                              summary_stats: Dict[str, Any],
                              output_dir: Path):
        """Create visualizations for the unified report"""
        try:
            from .visualization import TEIEvaluationVisualizer

            visualizer = TEIEvaluationVisualizer(str(output_dir))
            results = visualizer.create_all_visualizations(data, summary_stats, schema_mode=self.schema_mode)

            if results:
                print(f"[OUTPUT] Visualizations created:")
                for viz_name, viz_path in results.items():
                    if viz_path:
                        print(f"[OUTPUT]    - {viz_name}: {viz_path}")
            else:
                print("[WARNING] No visualizations were created")

        except ImportError as e:
            print(f"[WARNING] Visualization module not available: {e}")
            print("[WARNING]    Install matplotlib, pandas, seaborn for visualizations")
        except Exception as e:
            print(f"[ERROR] Error creating visualizations: {e}")
            import traceback
            traceback.print_exc()

    def _save_unified_csv(self, data: List[Dict[str, Any]], file_path: Path):
        """Save unified data as CSV (fallback method)"""
        if not data:
            return

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)

    def _save_unified_json(self, data: List[Dict[str, Any]], file_path: Path):
        """Save unified data as JSON with metadata"""
        report_data = {
            "timestamp": datetime.now().isoformat(),
            "report_type": "Unified TEI Evaluation Summary",
            "dimensions_included": ["Dimension 0 (Wellformedness)", "Dimension 1 (Content Preservation)", "Dimension 2 (Schema Validation)", "Dimension 3 (Structural Comparison)", "Dimension 4 (Content Matching)"],
            "total_files": len(data),
            "summary_statistics": self._generate_summary_stats(data),
            "files": data
        }

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)

    def _generate_summary_stats(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate comprehensive summary statistics across all dimensions"""
        if not data:
            return {}

        total_files = len(data)

        # Helper function to safely get numeric values
        def get_num_values(key):
            return [d[key] for d in data if d[key] != "N/A" and isinstance(d[key], (int, float))]
        # Helper: coerce to int when possible (handles numeric strings)
        def to_int_safe(val):
            if isinstance(val, int):
                return val
            if isinstance(val, float):
                try:
                    return int(val)
                except (ValueError, OverflowError):
                    return None
            if isinstance(val, str):
                try:
                    return int(val.strip())
                except (ValueError, AttributeError):
                    return None
            return None

        # ========== DIMENSION 1: WELLFORMEDNESS ==========
        # Note: Wellformedness metrics come from Dimension 0 evaluation keys (d0_*)
        d1_wellformedness_available = sum(1 for d in data if d.get("d0_well_formed", "N/A") != "N/A")
        d1_wellformedness_passed = sum(1 for d in data if d.get("d0_well_formed", "N/A") is True)
        d1_wellformedness_failed = sum(1 for d in data if d.get("d0_well_formed", "N/A") is False)

        d1_xml_scores = [d.get("d0_score") for d in data if isinstance(d.get("d0_score"), (int, float))]
        d1_total_errors = sum(d.get("d0_xml_errors", 0) for d in data if isinstance(d.get("d0_xml_errors", 0), int))
        d1_files_not_found = sum(1 for d in data if d.get("d0_file_not_found") is True)
        d1_non_xml_errors_total = sum(d.get("d0_non_xml_errors", 0) for d in data if isinstance(d.get("d0_non_xml_errors", 0), (int, float)))

        dimension1_wellformedness = {
            "total_files": d1_wellformedness_available,
            "passed_files": d1_wellformedness_passed,
            "failed_files": d1_wellformedness_failed,
            "pass_rate": (d1_wellformedness_passed / d1_wellformedness_available * 100) if d1_wellformedness_available > 0 else 0,
            "average_xml_score": sum(d1_xml_scores) / len(d1_xml_scores) if d1_xml_scores else 0,
            "min_xml_score": min(d1_xml_scores) if d1_xml_scores else 0,
            "max_xml_score": max(d1_xml_scores) if d1_xml_scores else 0,
            "total_errors": d1_total_errors,
            "files_not_found": d1_files_not_found,
            "non_xml_errors_total": d1_non_xml_errors_total
        }

        # ========== DIMENSION 1: CONTENT PRESERVATION ==========
        d1_content_available = sum(1 for d in data if d.get("d1_source_fidelity", "N/A") != "N/A")
        d1_content_passed = sum(1 for d in data if d.get("d1_source_fidelity") is True)
        d1_content_failed = sum(1 for d in data if d.get("d1_source_fidelity") is False)

        d1_content_scores = get_num_values("d1_source_fidelity_score")
        d1_total_differences = sum(d["d1_content_errors"] for d in data if isinstance(d["d1_content_errors"], int))

        dimension1_content = {
            "total_files": d1_content_available,
            "passed_files": d1_content_passed,
            "failed_files": d1_content_failed,
            "pass_rate": (d1_content_passed / d1_content_available * 100) if d1_content_available > 0 else 0,
            "average_content_score": sum(d1_content_scores) / len(d1_content_scores) if d1_content_scores else 0,
            "min_content_score": min(d1_content_scores) if d1_content_scores else 0,
            "max_content_score": max(d1_content_scores) if d1_content_scores else 0,
            "total_differences": d1_total_differences
        }

        # ========== DIMENSION 2: TEI VALIDATION ==========
        d2_tei_available = sum(1 for d in data if d["d2_tei_valid"] != "N/A")
        d2_tei_passed = sum(1 for d in data if d["d2_tei_valid"] is True)
        d2_tei_failed = sum(1 for d in data if d["d2_tei_valid"] is False)

        d2_tei_scores = get_num_values("d2_tei_score")

        dimension2_tei = {
            "total_files": d2_tei_available,
            "valid_files": d2_tei_passed,
            "invalid_files": d2_tei_failed,
            "pass_rate": (d2_tei_passed / d2_tei_available * 100) if d2_tei_available > 0 else 0,
            "average_tei_score": sum(d2_tei_scores) / len(d2_tei_scores) if d2_tei_scores else 0,
            "min_tei_score": min(d2_tei_scores) if d2_tei_scores else 0,
            "max_tei_score": max(d2_tei_scores) if d2_tei_scores else 0
        }

        # ========== DIMENSION 2: PROJECT VALIDATION ==========
        d2_project_available = sum(1 for d in data if d["d2_project_valid"] != "N/A")
        d2_project_passed = sum(1 for d in data if d["d2_project_valid"] is True)
        d2_project_failed = sum(1 for d in data if d["d2_project_valid"] is False)

        d2_project_scores = get_num_values("d2_project_score")

        dimension2_project = {
            "total_files": d2_project_available,
            "valid_files": d2_project_passed,
            "invalid_files": d2_project_failed,
            "pass_rate": (d2_project_passed / d2_project_available * 100) if d2_project_available > 0 else 0,
            "average_project_score": sum(d2_project_scores) / len(d2_project_scores) if d2_project_scores else 0,
            "min_project_score": min(d2_project_scores) if d2_project_scores else 0,
            "max_project_score": max(d2_project_scores) if d2_project_scores else 0
        }

        # ========== DIMENSION 3: STRUCTURAL COMPARISON ==========
        d3_available = sum(1 for d in data if d["d3_structural_match"] != "N/A")
        d3_structural_passed = sum(1 for d in data if d["d3_structural_match"] is True)
        d3_structural_failed = sum(1 for d in data if d["d3_structural_match"] is False)
        d3_structural_scores = get_num_values("d3_structural_score")

        dimension3_structural = {
            "total_files": d3_available,
            "matched_files": d3_structural_passed,
            "different_files": d3_structural_failed,
            "match_rate": (d3_structural_passed / d3_available * 100) if d3_available > 0 else 0,
            "average_structural_score": sum(d3_structural_scores) / len(d3_structural_scores) if d3_structural_scores else 0,
            "min_structural_score": min(d3_structural_scores) if d3_structural_scores else 0,
            "max_structural_score": max(d3_structural_scores) if d3_structural_scores else 0
        }

        # ========== DIMENSION 4: CONTENT MATCHING ==========
        d4_available = sum(1 for d in data if d["d4_content_match"] != "N/A")
        d4_content_passed = sum(1 for d in data if d["d4_content_match"] is True)
        d4_content_failed = sum(1 for d in data if d["d4_content_match"] is False)

        d4_scores = get_num_values("d4_semantic_score")

        # Aggregate exact and practical matches
        d4_exact_matches = sum(
            v for v in (to_int_safe(d.get("d4_exact_content_matches")) for d in data)
            if v is not None
        )
        d4_practical_matches = sum(
            v for v in (to_int_safe(d.get("d4_practical_content_matches")) for d in data)
            if v is not None
        )
        d4_total_elements = sum(
            v for v in (to_int_safe(d.get("d4_total_elements")) for d in data)
            if v is not None
        )

        # Calculate total comparisons from actual total_elements
        # This represents ALL compared elements (exact + practical + no-match)
        total_comparisons = d4_total_elements if d4_total_elements > 0 else (d4_exact_matches + d4_practical_matches)

        dimension4_content = {
            "total_files": d4_available,
            "passed_files": d4_content_passed,
            "failed_files": d4_content_failed,
            "pass_rate": (d4_content_passed / d4_available * 100) if d4_available > 0 else 0,
            "average_score": sum(d4_scores) / len(d4_scores) if d4_scores else 0,
            "min_score": min(d4_scores) if d4_scores else 0,
            "max_score": max(d4_scores) if d4_scores else 0,
            "content_stats": {
                "total_exact_matches": d4_exact_matches,
                "total_practical_matches": d4_practical_matches,
                "exact_match_rate": (d4_exact_matches / total_comparisons * 100) if total_comparisons > 0 else 0,
                "practical_match_rate": (d4_practical_matches / total_comparisons * 100) if total_comparisons > 0 else 0
            }
        }

        # ========== OVERALL TIERED SCORING ==========
        # Get overall scores - invalid files get 0 points and are included in the average
        # Use overall_score_all_files which includes 0.0 for invalid files
        overall_scores = []
        for d in data:
            score = d.get("overall_score_all_files", 0.0)
            if isinstance(score, (int, float)):
                overall_scores.append(score)
            else:
                # If somehow not numeric, default to 0.0 for invalid files
                overall_scores.append(0.0)
        overall_scores_all_files = overall_scores.copy()
        core_quality_scores = get_num_values("core_quality_score")
        d4_adjusted_scores_sem = get_num_values("d4_adjusted_semantic_score")

        # Count by evaluation status
        valid_files = sum(1 for d in data if d.get("evaluation_status") == "VALID")
        invalid_files = sum(1 for d in data if d.get("evaluation_status") == "INVALID")

        # Schema factor statistics
        schema_factors = [d.get("schema_factor") for d in data
                         if d.get("schema_factor") != "N/A" and isinstance(d.get("schema_factor"), (int, float))]

        overall_tiered_scoring = {
            "schema_mode": self.schema_mode,
            "total_files": total_files,
            "valid_files": valid_files,
            "invalid_files": invalid_files,
            "validity_rate": (valid_files / total_files * 100) if total_files > 0 else 0,
            "average_overall_score": sum(overall_scores) / len(overall_scores) if overall_scores else 0,
            "min_overall_score": min(overall_scores) if overall_scores else 0,
            "max_overall_score": max(overall_scores) if overall_scores else 0,
            "average_overall_score_all_files": sum(overall_scores_all_files) / len(overall_scores_all_files) if overall_scores_all_files else 0,
            "min_overall_score_all_files": min(overall_scores_all_files) if overall_scores_all_files else 0,
            "max_overall_score_all_files": max(overall_scores_all_files) if overall_scores_all_files else 0,
            "average_core_quality": sum(core_quality_scores) / len(core_quality_scores) if core_quality_scores else 0,
            "min_core_quality": min(core_quality_scores) if core_quality_scores else 0,
            "max_core_quality": max(core_quality_scores) if core_quality_scores else 0,
            "average_d4_adjusted_semantic": sum(d4_adjusted_scores_sem) / len(d4_adjusted_scores_sem) if d4_adjusted_scores_sem else 0,
            "average_schema_factor": sum(schema_factors) / len(schema_factors) if schema_factors else 1.0,
            "min_schema_factor": min(schema_factors) if schema_factors else 1.0,
            "max_schema_factor": max(schema_factors) if schema_factors else 1.0
        }

        return {
            "total_files": total_files,
            "dimension1_wellformedness": dimension1_wellformedness,
            "dimension1_content": dimension1_content,
            "dimension2_tei": dimension2_tei,
            "dimension2_project": dimension2_project,
            "dimension3_structural": dimension3_structural,
            "dimension4_content": dimension4_content,
            "overall_tiered_scoring": overall_tiered_scoring
        }

def create_unified_report(evaluation_output_dir: str = "evaluation_output", schema_mode: str = "both"):
    """
    Convenience function to create unified report from standard evaluation output structure

    Args:
        evaluation_output_dir: Base directory containing dimension evaluation results
        schema_mode: Schema validation mode for overall scoring
            - "none": No schema validation impact on score
            - "tei": TEI schema validation only
            - "project": Project schema validation only
            - "both": Both TEI and project schema validation (default)
    """
    base_path = Path(evaluation_output_dir)

    if not base_path.exists():
        print(f"[ERROR] Evaluation output directory not found: {evaluation_output_dir}")
        return

    reporter = UnifiedReporter(schema_mode=schema_mode)

    print(f"[INFO] Schema validation mode: {schema_mode}")

    # Load results from base path (new flat structure)
    print(f"[OUTPUT] Loading results from: {base_path}")
    reporter.load_dimension0_results(str(base_path))
    reporter.load_dimension1_results(str(base_path))
    reporter.load_dimension2_results(str(base_path))
    reporter.load_dimension3_results(str(base_path))
    reporter.load_dimension4_results(str(base_path))

    # Generate unified report directly in the processing directory (root/json structure)
    reporter.generate_unified_report(str(base_path))

if __name__ == "__main__":
    # Example usage
    create_unified_report("evaluation_output")